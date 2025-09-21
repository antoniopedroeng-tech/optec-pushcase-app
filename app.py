import os
import io
import csv
import traceback
from datetime import datetime, date, timedelta
from flask import Flask, render_template, render_template_string, request, redirect, url_for, session, flash, send_file
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from werkzeug.exceptions import HTTPException

APP_NAME = "OPTEC PUSHCASE APP"
SECRET_KEY = os.environ.get("SECRET_KEY", "dev-secret-change-me")
DATABASE_URL = os.environ.get("DATABASE_URL")  # fornecido pelo Render Postgres
TIMEZONE_TZ = os.environ.get("TZ", "America/Fortaleza")

# ============================ ENGINE / SESSION ============================
# Observação: se DATABASE_URL estiver vazio/None, create_engine quebra.
if not DATABASE_URL:
    print("[BOOT] ATENÇÃO: DATABASE_URL não está definido! Defina a variável de ambiente.", flush=True)

engine = create_engine(DATABASE_URL, pool_pre_ping=True) if DATABASE_URL else None
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False) if engine else None

app = Flask(__name__)
app.secret_key = SECRET_KEY

# ============================ ERROS / DIAGNÓSTICO ============================

@app.errorhandler(Exception)
def handle_exception(e):
    if isinstance(e, HTTPException):
        return e
    tb = traceback.format_exc()
    print(f"[ERROR] Rota: {request.path}\n{tb}", flush=True)
    html = """
    {% extends "base.html" %}
    {% block title %}Erro Interno (500){% endblock %}
    {% block content %}
      <h2>Erro Interno (500)</h2>
      <p>Ocorreu um erro ao processar <code>{{ path }}</code>.</p>
      <p>Tente novamente. O administrador pode verificar os logs do servidor para mais detalhes.</p>
      <details style="margin-top:12px;">
        <summary>Mostrar detalhes técnicos (stack trace)</summary>
        <pre style="white-space:pre-wrap;background:#f7f7f7;border:1px solid #ddd;padding:8px;border-radius:8px;">{{ tb }}</pre>
      </details>
    {% endblock %}
    """
    try:
        return render_template_string(html, path=request.path, tb=tb), 500
    except Exception:
        return (f"<h1>500 - Erro Interno</h1><p>Rota: {request.path}</p>"
                "<p>Veja os logs do servidor para detalhes.</p>"), 500

@app.route("/_diag")
def diag():
    info = {
        "app_name": APP_NAME,
        "now_utc": datetime.utcnow().isoformat(sep=" ", timespec="seconds"),
        "database_url_prefix": (DATABASE_URL.split("://")[0] + "://***") if DATABASE_URL else "(vazio)",
        "engine_ready": bool(engine),
        "openpyxl": None,
        "db_ok": None,
        "tables": [],
        "users_count": None,
    }
    try:
        import openpyxl  # noqa
        info["openpyxl"] = "instalado"
    except Exception as e:
        info["openpyxl"] = f"ausente ({e})"

    if engine:
        try:
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
                info["db_ok"] = True
                rs = conn.execute(text("""
                    SELECT tablename FROM pg_tables
                    WHERE schemaname NOT IN ('pg_catalog','information_schema')
                    ORDER BY tablename
                """))
                info["tables"] = [r[0] for r in rs.fetchall()]
                try:
                    cnt = conn.execute(text("SELECT COUNT(*) FROM users")).scalar_one()
                    info["users_count"] = int(cnt)
                except Exception as e:
                    info["users_count"] = f"falhou ({e})"
        except Exception as e:
            info["db_ok"] = f"falhou ({e})"
    else:
        info["db_ok"] = "engine não inicializado (DATABASE_URL?)"

    html = """
    <!DOCTYPE html><html lang="pt-BR"><head><meta charset="utf-8"><title>Diagnóstico</title>
    <style>
      body{font-family:Arial,sans-serif;max-width:880px;margin:24px auto;padding:0 12px;}
      code,pre{background:#f7f7f7;border:1px solid #ddd;padding:4px 6px;border-radius:6px;}
      table{border-collapse:collapse;width:100%;margin-top:10px;}
      th,td{border:1px solid #ddd;padding:8px;text-align:left;}
      th{background:#eee;}
    </style></head><body>
    <h2>Diagnóstico</h2>
    <ul>
      <li><b>App:</b> {{ info.app_name }}</li>
      <li><b>Agora (UTC):</b> {{ info.now_utc }}</li>
      <li><b>DATABASE_URL:</b> {{ info.database_url_prefix }}</li>
      <li><b>Engine pronto:</b> {{ info.engine_ready }}</li>
      <li><b>Conexão DB:</b> {{ info.db_ok }}</li>
      <li><b>openpyxl:</b> {{ info.openpyxl }}</li>
      <li><b>Users count:</b> {{ info.users_count }}</li>
    </ul>
    <h3>Tabelas</h3>
    {% if info.tables %}
      <ul>{% for t in info.tables %}<li><code>{{ t }}</code></li>{% endfor %}</ul>
    {% else %}
      <p>(Nenhuma tabela retornada.)</p>
    {% endif %}
    </body></html>
    """
    return render_template_string(html, info=info)

# ============================ DB INIT ============================

def init_db():
    if not engine:
        print("[BOOT] init_db(): pulado porque engine não foi criado (DATABASE_URL ausente).", flush=True)
        return

    ddl = """
    CREATE TABLE IF NOT EXISTS users (
      id SERIAL PRIMARY KEY,
      username TEXT UNIQUE NOT NULL,
      password_hash TEXT NOT NULL,
      role TEXT NOT NULL CHECK (role IN ('admin','comprador','pagador','cliente')),
      created_at TIMESTAMP NOT NULL
    );

    CREATE TABLE IF NOT EXISTS suppliers (
      id SERIAL PRIMARY KEY,
      name TEXT UNIQUE NOT NULL,
      active INTEGER NOT NULL DEFAULT 1,
      billing INTEGER NOT NULL DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS products (
      id SERIAL PRIMARY KEY,
      name TEXT NOT NULL,
      code TEXT,
      kind TEXT NOT NULL,
      active INTEGER NOT NULL DEFAULT 1,
      in_stock INTEGER NOT NULL DEFAULT 0,
      UNIQUE(name, kind)
    );

    CREATE TABLE IF NOT EXISTS rules (
      id SERIAL PRIMARY KEY,
      product_id INTEGER NOT NULL REFERENCES products(id),
      supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
      max_price DOUBLE PRECISION NOT NULL,
      active INTEGER NOT NULL DEFAULT 1,
      UNIQUE(product_id, supplier_id)
    );

    CREATE TABLE IF NOT EXISTS purchase_orders (
      id SERIAL PRIMARY KEY,
      buyer_id INTEGER NOT NULL REFERENCES users(id),
      supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
      status TEXT NOT NULL CHECK (status IN ('PENDENTE_PAGAMENTO','PAGO','CANCELADO')),
      total DOUBLE PRECISION NOT NULL,
      note TEXT,
      created_at TIMESTAMP NOT NULL,
      updated_at TIMESTAMP NOT NULL
    );

    CREATE TABLE IF NOT EXISTS purchase_items (
      id SERIAL PRIMARY KEY,
      order_id INTEGER NOT NULL REFERENCES purchase_orders(id) ON DELETE CASCADE,
      product_id INTEGER NOT NULL REFERENCES products(id),
      quantity INTEGER NOT NULL,
      unit_price DOUBLE PRECISION NOT NULL,
      sphere DOUBLE PRECISION,
      cylinder DOUBLE PRECISION,
      base DOUBLE PRECISION,
      addition DOUBLE PRECISION,
      os_number TEXT
    );

    DO $$ BEGIN
      IF EXISTS (SELECT 1 FROM pg_indexes WHERE indexname = 'idx_purchase_items_os') THEN
        EXECUTE 'DROP INDEX idx_purchase_items_os';
      END IF;
    EXCEPTION WHEN others THEN
      NULL;
    END $$;

    CREATE TABLE IF NOT EXISTS payments (
      id SERIAL PRIMARY KEY,
      order_id INTEGER NOT NULL UNIQUE REFERENCES purchase_orders(id) ON DELETE CASCADE,
      payer_id INTEGER NOT NULL REFERENCES users(id),
      method TEXT,
      reference TEXT,
      paid_at TIMESTAMP NOT NULL,
      amount DOUBLE PRECISION NOT NULL
    );

    
    CREATE TABLE IF NOT EXISTS supplier_credits (
      id SERIAL PRIMARY KEY,
      supplier_id INTEGER NOT NULL REFERENCES suppliers(id),
      item_id INTEGER NOT NULL UNIQUE REFERENCES purchase_items(id) ON DELETE CASCADE,
      amount DOUBLE PRECISION NOT NULL,
      created_at TIMESTAMP NOT NULL
    );
CREATE TABLE IF NOT EXISTS audit_log (
      id SERIAL PRIMARY KEY,
      user_id INTEGER REFERENCES users(id),
      action TEXT NOT NULL,
      details TEXT,
      created_at TIMESTAMP NOT NULL
    );
    
    -- Orçamento: tabelas
    CREATE TABLE IF NOT EXISTS orc_produto (
      id SERIAL PRIMARY KEY,
      code TEXT UNIQUE NOT NULL,
      name TEXT NOT NULL,
      price NUMERIC(12,2) NOT NULL DEFAULT 0,
      visao TEXT NOT NULL CHECK (visao IN ('visao_simples','progressiva','bifocal')),
      ar INTEGER NOT NULL DEFAULT 0,
      foto INTEGER NOT NULL DEFAULT 0,
      azul INTEGER NOT NULL DEFAULT 0,
      esf_min NUMERIC(6,2) NOT NULL DEFAULT 0,
      esf_max NUMERIC(6,2) NOT NULL DEFAULT 0,
      cil_min NUMERIC(6,2) NOT NULL DEFAULT 0,
      cil_max NUMERIC(6,2) NOT NULL DEFAULT 0,
      updated_at TIMESTAMP NOT NULL DEFAULT NOW()
    );
    CREATE TABLE IF NOT EXISTS orc_servico_catalogo (
      code TEXT PRIMARY KEY,
      description TEXT,
      price NUMERIC(12,2) NOT NULL DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS orc_produto_serv_obrig (
      id SERIAL PRIMARY KEY,
      produto_id INTEGER NOT NULL REFERENCES orc_produto(id) ON DELETE CASCADE,
      serv_code TEXT NOT NULL REFERENCES orc_servico_catalogo(code) ON DELETE RESTRICT,
      UNIQUE (produto_id, serv_code)
    );
    CREATE TABLE IF NOT EXISTS orc_produto_serv_opc (
      id SERIAL PRIMARY KEY,
      produto_id INTEGER NOT NULL REFERENCES orc_produto(id) ON DELETE CASCADE,
      serv_code TEXT NOT NULL REFERENCES orc_servico_catalogo(code) ON DELETE RESTRICT,
      UNIQUE (produto_id, serv_code)
    );
    CREATE TABLE IF NOT EXISTS orc_produto_acrescimo (
      id SERIAL PRIMARY KEY,
      produto_id INTEGER NOT NULL REFERENCES orc_produto(id) ON DELETE CASCADE,
      serv_code TEXT NOT NULL REFERENCES orc_servico_catalogo(code) ON DELETE RESTRICT,
      esf_min NUMERIC(6,2), esf_max NUMERIC(6,2),
      cil_min NUMERIC(6,2), cil_max NUMERIC(6,2),
      UNIQUE (produto_id, serv_code, esf_min, esf_max, cil_min, cil_max)
    );
"""
    with engine.begin() as conn:
        conn.execute(text(ddl))
        try:
            conn.execute(text("ALTER TABLE products ADD COLUMN IF NOT EXISTS in_stock INTEGER NOT NULL DEFAULT 0"))
        except Exception:
            pass
        try:
            conn.execute(text("ALTER TABLE suppliers ADD COLUMN IF NOT EXISTS billing INTEGER NOT NULL DEFAULT 0"))
        except Exception:
            pass

        exists = conn.execute(text("SELECT COUNT(*) AS n FROM users")).scalar_one()
        if exists == 0:
            from werkzeug.security import generate_password_hash
            conn.execute(
                text("INSERT INTO users (username, password_hash, role, created_at) VALUES (:u,:p,:r,:c)"),
                dict(u="admin", p=generate_password_hash("admin123"), r="admin", c=datetime.utcnow())
            )

# Helpers comuns
def db_all(sql, **params):
    with engine.connect() as conn:
        return conn.execute(text(sql), params).mappings().all()

def db_one(sql, **params):
    with engine.connect() as conn:
        return conn.execute(text(sql), params).mappings().first()

def db_exec(sql, **params):
    with engine.begin() as conn:
        conn.execute(text(sql), params)

def audit(action, details=""):
    u = current_user()
    db_exec("INSERT INTO audit_log (user_id, action, details, created_at) VALUES (:uid,:a,:d,:c)",
            uid=(u["id"] if u else None), a=action, d=details, c=datetime.utcnow())

# ============ NOVO: Troca automática por cilindro (CIL./CIL. EST./CIL. SUPER EST.) ============
def maybe_swap_lente_by_cylinder(product_id: int, supplier_id: int, current_price: float, cylinder: float):
    """
    Regras:
      - -2.25 até -4.00  → troca para '<nome> CIL. EST.' (fallback: ' CIL.')
      - menor que -4.00  → troca para '<nome> CIL. SUPER EST.' (fallback: ' CIL. EST.' → ' CIL.')
    Retorna (new_product_id, new_price, (old_name, new_name)) ou (product_id, current_price, None).
    Ajusta preço para o teto da regra do novo produto quando necessário.
    """
    if cylinder is None:
        return product_id, current_price, None

    base = db_one("SELECT id, name, kind FROM products WHERE id=:id", id=product_id)
    if not base or (base["kind"] or "") != "lente":
        return product_id, current_price, None

    cyl = float(cylinder)
    if -4.00 <= cyl <= -2.25:
        suffixes = [" CIL. EST.", " CIL."]
    elif cyl < -4.00:
        suffixes = [" CIL. SUPER EST.", " CIL. EST.", " CIL."]
    else:
        return product_id, current_price, None

    for sfx in suffixes:
        cand = db_one(
            "SELECT id, name FROM products WHERE name=:n AND kind='lente' AND active=1",
            n=(base["name"] + sfx)
        )
        if not cand:
            continue

        rule = db_one(
            "SELECT max_price FROM rules WHERE product_id=:pid AND supplier_id=:sid AND active=1",
            pid=cand["id"], sid=supplier_id
        )
        if not rule:
            continue

        new_price = float(current_price or 0)
        maxp = float(rule["max_price"])
        if new_price <= 0 or new_price > maxp + 1e-6:
            new_price = maxp

        return cand["id"], new_price, (base["name"], cand["name"])

    return product_id, current_price, None

# ============================ AUTH/CTX ============================

def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    u = db_one("SELECT * FROM users WHERE id=:id", id=uid)
    return u

def require_role(*roles):
    u = current_user()
    if not u or u["role"] not in roles:
        flash("Acesso negado.", "error")
        return redirect(url_for("index"))

@app.context_processor
def inject_globals():
    return {"now": datetime.utcnow(), "role": session.get("role"), "user": current_user(), "app_name": APP_NAME}

# ============================ RELATÓRIOS (Excel in-memory) ============================

def build_excel_bytes_for_day(day_str: str) -> bytes:
    rows = db_all("""
        SELECT
            s.name  AS fornecedor,
            p.name  AS produto,
            p.in_stock AS in_stock,
            i.sphere, i.cylinder, i.base, i.addition,
            i.quantity, i.unit_price,
            i.os_number AS os,
            pay.method AS metodo,
            DATE(pay.paid_at) AS data
        FROM payments pay
        JOIN purchase_orders o ON o.id = pay.order_id
        JOIN suppliers s       ON s.id = o.supplier_id
        JOIN purchase_items i  ON i.order_id = o.id
        JOIN products p        ON p.id = i.product_id
        WHERE DATE(pay.paid_at) = :day
        ORDER BY s.name, p.name
    """, day=day_str)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("openpyxl não está instalado") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos do Dia"
    ws.append(["Fornecedor", "OS", "Produto", "Estoque", "Dioptria", "Data", "Método", "Valor"])

    def fmt_dioptria(r):
        if r["sphere"] is not None or r["cylinder"] is not None:
            esf = f"{r['sphere']:+.2f}" if r["sphere"] is not None else "-"
            cil = f"{r['cylinder']:+.2f}" if r["cylinder"] is not None else "-"
            return f"Esf {esf} / Cil {cil}"
        else:
            b = f"{r['base']:.2f}" if r["base"] is not None else "-"
            add = f"+{r['addition']:.2f}" if r["addition"] is not None else "-"
            return f"Base {b} / Adição {add}"

    grand_total = 0.0
    for r in rows:
        subtotal = float(r["quantity"] or 0) * float(r["unit_price"] or 0.0)
        grand_total += subtotal
        ws.append([
            r["fornecedor"],
            r["os"] or "",
            r["produto"],
            "Sim" if int(r["in_stock"] or 0) == 1 else "Não",
            fmt_dioptria(r),
            r["data"].isoformat() if hasattr(r["data"], "isoformat") else str(r["data"]),
            r["metodo"] or "",
            float(f"{subtotal:.2f}")
        ])

    ws.append(["", "", "", "", "", "", ""])
    from openpyxl.styles import Font
    ws.append(["", "", "", "", "", "TOTAL", float(f"{grand_total:.2f}")])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)

    from openpyxl.utils import get_column_letter
    for i, w in enumerate([18, 28, 12, 26, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def build_excel_bytes_for_period(start_str: str, end_str: str) -> bytes:
    rows = db_all("""
        SELECT
            s.name  AS fornecedor,
            p.name  AS produto,
            p.in_stock AS in_stock,
            i.sphere, i.cylinder, i.base, i.addition,
            i.quantity, i.unit_price,
            i.os_number AS os,
            pay.method AS metodo,
            DATE(pay.paid_at) AS data
        FROM payments pay
        JOIN purchase_orders o ON o.id = pay.order_id
        JOIN suppliers s       ON s.id = o.supplier_id
        JOIN purchase_items i  ON i.order_id = o.id
        JOIN products p        ON p.id = i.product_id
        WHERE DATE(pay.paid_at) BETWEEN :d1 AND :d2
        ORDER BY DATE(pay.paid_at), s.name, p.name
    """, d1=start_str, d2=end_str)

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
    except ImportError as e:
        raise RuntimeError("openpyxl não está instalado") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos por Período"
    ws.append(["Fornecedor", "OS", "Produto", "Estoque", "Dioptria", "Data", "Método", "Valor"])

    def fmt_dioptria(r):
        if r["sphere"] is not None or r["cylinder"] is not None:
            esf = f"{r['sphere']:+.2f}" if r["sphere"] is not None else "-"
            cil = f"{r['cylinder']:+.2f}" if r["cylinder"] is not None else "-"
            return f"Esf {esf} / Cil {cil}"
        else:
            b = f"{r['base']:.2f}" if r["base"] is not None else "-"
            add = f"+{r['addition']:.2f}" if r["addition"] is not None else "-"
            return f"Base {b} / Adição {add}"

    grand_total = 0.0
    for r in rows:
        subtotal = float(r["quantity"] or 0) * float(r["unit_price"] or 0.0)
        grand_total += subtotal
        ws.append([
            r["fornecedor"],
            r["os"] or "",
            r["produto"],
            "Sim" if int(r["in_stock"] or 0) == 1 else "Não",
            fmt_dioptria(r),
            r["data"].isoformat() if hasattr(r["data"], "isoformat") else str(r["data"]),
            r["metodo"] or "",
            float(f"{subtotal:.2f}")
        ])

    ws.append(["", "", "", "", "", "", ""])
    ws.append(["", "", "", "", "", "TOTAL", float(f"{grand_total:.2f}")])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)

    from openpyxl.utils import get_column_letter
    for i, w in enumerate([18, 28, 12, 26, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ============================ ROTAS ============================

# --- Tela Orçamento ----------------------------------------------------------
@app.route("/orcamento", methods=["GET"])
def orcamento():
    # permite admin, comprador, pagador e cliente
    ret = require_role("admin", "comprador", "pagador", "cliente")
    if ret:
        return ret
    return render_template("orcamento.html")


@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        from werkzeug.security import check_password_hash
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        u = db_one("SELECT * FROM users WHERE username=:u", u=username)
        if u and check_password_hash(u["password_hash"], password):
            session["user_id"] = u["id"]; session["role"] = u["role"]
            flash(f"Bem-vindo, {u['username']}!", "success"); audit("login", f"user={u['username']}")
            return redirect(url_for("index"))
        flash("Credenciais inválidas", "error")
    try:
        return render_template("login.html")
    except Exception:
        return render_template_string("""
        <!doctype html><meta charset="utf-8"><title>Login</title>
        <h2>Login</h2>
        <form method="post">
          <div><label>Usuário</label><br><input name="username" required></div>
          <div><label>Senha</label><br><input name="password" type="password" required></div>
          <button>Entrar</button>
        </form>
        """)

@app.route("/logout")
def logout():
    u = current_user(); session.clear(); flash("Sessão encerrada.", "info"); audit("logout", f"user={u['username'] if u else ''}")
    return redirect(url_for("login"))

@app.route("/")
def index():
    try:
        return render_template("index.html")
    except Exception:
        return render_template_string("""
        {% extends "base.html" %}
        {% block content %}
          <h2>Bem-vindo ao {{ app_name }}</h2>
          <p>Use o menu superior para navegar.</p>
        {% endblock %}
        """, app_name=APP_NAME)

# ==================== APIs da tela Orçamento ====================
from decimal import Decimal
from sqlalchemy import text
from flask import request

def _dec(v, default="0"):
    try:
        if v in (None, ""): return Decimal(default)
        return Decimal(str(v).replace(",", "."))
    except Exception:
        return Decimal(default)

@app.post("/api/orcamento/options")
def api_orcamento_options():
    # permite admin, comprador, pagador e cliente
    ret = require_role("admin", "comprador", "pagador", "cliente")
    if ret:
        return ret

    data = request.get_json(force=True) or {}
    visao = (data.get("visao") or "").strip()
    flags = data.get("flags") or {}
    ar   = bool(flags.get("ar"))
    ar_i = 1 if ar else 0
    foto = bool(flags.get("foto"))
    foto_i = 1 if foto else 0
    azul = bool(flags.get("azul"))
    azul_i = 1 if azul else 0

    od = data.get("od") or {}; oe = data.get("oe") or {}
    od_esf = _dec(od.get("esf")); od_cil = _dec(od.get("cil"))
    oe_esf = _dec(oe.get("esf")); oe_cil = _dec(oe.get("cil"))

    q = text("""
      SELECT id, name, price, esf_min, esf_max, cil_min, cil_max
      FROM orc_produto
      WHERE visao=:visao AND ar=:ar AND foto=:foto AND azul=:azul
    """)

    def inrng(v, a, b):
        a = Decimal(a); b = Decimal(b)
        if a == 0 and b == 0:
            return True
        lo = min(a,b); hi = max(a,b)
        return Decimal(v) >= lo and Decimal(v) <= hi

    products = []
    with engine.begin() as conn:
        for r in conn.execute(q, {"visao":visao,"ar":ar_i,"foto":foto_i,"azul":azul_i}).mappings():
            ok_od = inrng(od_esf, r["esf_min"], r["esf_max"]) and inrng(od_cil, r["cil_min"], r["cil_max"])
            ok_oe = inrng(oe_esf, r["esf_min"], r["esf_max"]) and inrng(oe_cil, r["cil_min"], r["cil_max"])
            if ok_od and ok_oe:
                products.append({"id": int(r["id"]), "name": r["name"], "price": float(r["price"] or 0)})

    return {"products": products}

@app.post("/api/orcamento/services")
def api_orcamento_services():
    ret = require_role("admin","comprador","pagador","cliente")
    if ret:
        return ret

    data = request.get_json(force=True) or {}
    pid = int(data.get("product_id"))

    od = data.get("od") or {}; oe = data.get("oe") or {}
    od_esf = _dec(od.get("esf")); od_cil = _dec(od.get("cil"))
    oe_esf = _dec(oe.get("esf")); oe_cil = _dec(oe.get("cil"))

    with engine.begin() as conn:
        ob = conn.execute(text("""
          SELECT c.code, COALESCE(c.description,c.code) AS name, COALESCE(c.price,0) AS price
          FROM orc_produto_serv_obrig o
          JOIN orc_servico_catalogo c ON c.code=o.serv_code
          WHERE o.produto_id=:pid
        """), {"pid": pid}).mappings().all()

        op = conn.execute(text("""
          SELECT c.code, COALESCE(c.description,c.code) AS name, COALESCE(c.price,0) AS price
          FROM orc_produto_serv_opc o
          JOIN orc_servico_catalogo c ON c.code=o.serv_code
          WHERE o.produto_id=:pid
        """), {"pid": pid}).mappings().all()

        ac = conn.execute(text("""
          SELECT a.serv_code, a.esf_min, a.esf_max, a.cil_min, a.cil_max,
                 COALESCE(c.description,a.serv_code) AS name, COALESCE(c.price,0) AS price
          FROM orc_produto_acrescimo a
          JOIN orc_servico_catalogo c ON c.code=a.serv_code
          WHERE a.produto_id=:pid
        """), {"pid": pid}).mappings().all()

    def within(v, a, b):
        if a is None and b is None: return False
        a = Decimal(a) if a is not None else Decimal(v)
        b = Decimal(b) if b is not None else Decimal(v)
        lo = min(a,b); hi = max(a,b)
        return Decimal(v) >= lo and Decimal(v) <= hi

    mandatory = [{"id": r["code"], "code": r["code"], "name": r["name"], "display": r["name"], "price": float(r["price"])} for r in ob]

    for a in ac:
        esf_od_ok = within(od_esf, a["esf_min"], a["esf_max"])
        cil_od_ok = within(od_cil, a["cil_min"], a["cil_max"])
        esf_oe_ok = within(oe_esf, a["esf_min"], a["esf_max"])
        cil_oe_ok = within(oe_cil, a["cil_min"], a["cil_max"])
        trig_od = (esf_od_ok or cil_od_ok)
        trig_oe = (esf_oe_ok or cil_oe_ok)
        if trig_od or trig_oe:
            mandatory.append({"id": a["serv_code"], "code": a["serv_code"], "name": a["name"], "display": a["name"], "price": float(a["price"])})
    optional = [{"id": r["code"], "code": r["code"], "name": r["name"], "display": r["name"], "price": float(r["price"])} for r in op]
    return {"mandatory": mandatory, "optional": optional}
# ================== fim das APIs de Orçamento ==================



# -------- Admin: Usuários --------

@app.route("/admin/users")
def admin_users():
    ret = require_role("admin")
    if ret: return ret
    users = db_all("SELECT id, username, role, created_at FROM users ORDER BY id")
    return render_template("admin_users.html", users=users)

@app.route("/admin/users/create", methods=["POST"])
def admin_users_create():
    # Somente admin pode criar
    ret = require_role("admin")
    if ret:
        return ret

    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    role = request.form.get("role") or "comprador"

    # validações básicas
    if not username or not password or (role not in ("admin","comprador","pagador","cliente")):
        flash("Dados inválidos.", "error")
        return redirect(url_for("admin_users"))

    # pré-checagem: já existe? (case-insensitive)
    exists = db_one("SELECT 1 FROM users WHERE LOWER(username)=LOWER(:u)", u=username)
    if exists:
        flash("Usuário já existe.", "error")
        return redirect(url_for("admin_users"))

    from werkzeug.security import generate_password_hash
    from sqlalchemy.exc import IntegrityError

    try:
        db_exec(
            "INSERT INTO users (username, password_hash, role, created_at) "
            "VALUES (:u,:p,:r,:c)",
            u=username,
            p=generate_password_hash(password),
            r=role,
            c=datetime.utcnow()
        )
        audit("user_create", f"{username}/{role}")
        flash("Usuário criado.", "success")
    except IntegrityError as e:
        try:
            code = getattr(getattr(e, 'orig', None), 'pgcode', None)
            if code == '23505':
                flash("Usuário já existe.", "error")
            else:
                flash(f"Erro no banco (IntegrityError): {e}", "error")
        except Exception:
            flash("Usuário já existe.", "error")
    except Exception as e:
        flash(f"Erro ao criar usuário: {e}", "error")

    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:uid>/delete", methods=["POST"])
def admin_users_delete(uid):
    ret = require_role("admin")
    if ret: return ret
    if uid == session.get("user_id"):
        flash("Não é possível excluir o próprio usuário logado.", "error")
        return redirect(url_for("admin_users"))

    # Exclusão direta sem bloquear por referências (histórico permanece)
    try:
        db_exec("DELETE FROM users WHERE id=:id", id=uid)
        audit("user_delete", f"id={uid}")
        flash("Usuário removido.", "success")
    except Exception as e:
        flash(f"Falha ao excluir usuário: {e}", "error")
    return redirect(url_for("admin_users"))

# -------- Admin: Fornecedores --------

@app.route("/admin/suppliers")
def admin_suppliers():
    ret = require_role("admin")
    if ret: return ret
    suppliers = db_all("SELECT * FROM suppliers ORDER BY name")
    return render_template("admin_suppliers.html", suppliers=suppliers)

@app.route("/admin/suppliers/create", methods=["POST"])
def admin_suppliers_create():
    ret = require_role("admin")
    if ret: return ret
    name = (request.form.get("name") or "").strip()
    billing = 1 if (request.form.get("billing") in ("on","1","true","True")) else 0
    if not name:
        flash("Nome inválido.", "error"); return redirect(url_for("admin_suppliers"))
    try:
        db_exec("INSERT INTO suppliers (name, active, billing) VALUES (:n,1,:b)", n=name, b=billing)
        audit("supplier_create", f"{name} billing={billing}"); flash("Fornecedor criado.", "success")
    except Exception:
        flash("Fornecedor já existe.", "error")
    return redirect(url_for("admin_suppliers"))

@app.route("/admin/suppliers/<int:sid>/toggle", methods=["POST"])
def admin_suppliers_toggle(sid):
    ret = require_role("admin")
    if ret: return ret
    s = db_one("SELECT * FROM suppliers WHERE id=:id", id=sid)
    if not s: flash("Fornecedor não encontrado.", "error"); return redirect(url_for("admin_suppliers"))
    new_active = 0 if s["active"] else 1
    db_exec("UPDATE suppliers SET active=:a WHERE id=:id", a=new_active, id=sid)
    audit("supplier_toggle", f"id={sid} active={new_active}")
    return redirect(url_for("admin_suppliers"))

@app.route("/admin/suppliers/<int:sid>/toggle-billing", methods=["POST"])
def admin_suppliers_toggle_billing(sid):
    ret = require_role("admin")
    if ret: return ret
    s = db_one("SELECT * FROM suppliers WHERE id=:id", id=sid)
    if not s:
        flash("Fornecedor não encontrado.", "error"); return redirect(url_for("admin_suppliers"))
    new_billing = 0 if (s["billing"] or 0) == 1 else 1
    db_exec("UPDATE suppliers SET billing=:b WHERE id=:id", b=new_billing, id=sid)
    audit("supplier_toggle_billing", f"id={sid} billing={new_billing}")
    return redirect(url_for("admin_suppliers"))

@app.route("/admin/suppliers/<int:sid>/delete", methods=["POST"])
def admin_suppliers_delete(sid):
    ret = require_role("admin")
    if ret: return ret
    used_rule = db_one("SELECT 1 FROM rules WHERE supplier_id=:id LIMIT 1", id=sid)
    used_order = db_one("SELECT 1 FROM purchase_orders WHERE supplier_id=:id LIMIT 1", id=sid)
    if used_rule or used_order:
        flash("Não é possível excluir: fornecedor em uso (regras ou pedidos).", "error")
        return redirect(url_for("admin_suppliers"))
    db_exec("DELETE FROM suppliers WHERE id=:id", id=sid)
    audit("supplier_delete", f"id={sid}")
    flash("Fornecedor excluído.", "success")
    return redirect(url_for("admin_suppliers"))

# -------- Admin: Produtos --------

@app.route("/admin/products")
def admin_products():
    ret = require_role("admin")
    if ret: return ret
    products = db_all("SELECT * FROM products ORDER BY kind, name")
    return render_template("admin_products.html", products=products)

@app.route("/admin/products/create", methods=["POST"])
def admin_products_create():
    ret = require_role("admin")
    if ret: return ret
    name = (request.form.get("name") or "").strip()
    code = (request.form.get("code") or "").strip()
    kind = (request.form.get("kind") or "lente").lower()
    in_stock = 1 if (request.form.get("in_stock") in ("on","1","true","True")) else 0
    if kind not in ("lente","bloco") or not name:
        flash("Dados inválidos.", "error"); return redirect(url_for("admin_products"))
    try:
        db_exec(
            "INSERT INTO products (name, code, kind, in_stock, active) "
            "VALUES (:n,:c,:k,:instock,1)",
            n=name, c=code, k=kind, instock=in_stock
        )
        audit("product_create", f"{name}/{kind}/in_stock={in_stock}"); flash("Produto criado.", "success")
    except Exception:
        flash("Produto já existe para este tipo.", "error")
    return redirect(url_for("admin_products"))

@app.route("/admin/products/<int:pid>/toggle", methods=["POST"])
def admin_products_toggle(pid):
    ret = require_role("admin")
    if ret: return ret
    p = db_one("SELECT * FROM products WHERE id=:id", id=pid)
    if not p: flash("Produto não encontrado.", "error"); return redirect(url_for("admin_products"))
    new_active = 0 if p["active"] else 1
    db_exec("UPDATE products SET active=:a WHERE id=:id", a=new_active, id=pid)
    audit("product_toggle", f"id={pid} active={new_active}")
    return redirect(url_for("admin_products"))

@app.route("/admin/products/<int:pid>/delete", methods=["POST"])
def admin_products_delete(pid):
    ret = require_role("admin")
    if ret: return ret
    used_rule = db_one("SELECT 1 FROM rules WHERE product_id=:id LIMIT 1", id=pid)
    used_item = db_one("SELECT 1 FROM purchase_items WHERE product_id=:id LIMIT 1", id=pid)
    if used_rule or used_item:
        flash("Não é possível excluir: produto em uso (regras ou pedidos).", "error")
        return redirect(url_for("admin_products"))
    db_exec("DELETE FROM products WHERE id=:id", id=pid)
    audit("product_delete", f"id={pid}")
    flash("Produto excluído.", "success")
    return redirect(url_for("admin_products"))

# -------- Admin: Regras --------

@app.route("/admin/rules")
def admin_rules():
    ret = require_role("admin")
    if ret: return ret
    rules = db_all("""
        SELECT r.id, r.max_price, r.active,
               p.name as product_name, p.kind as product_kind, p.id as product_id,
               s.name as supplier_name, s.id as supplier_id
        FROM rules r
        JOIN products p ON p.id = r.product_id
        JOIN suppliers s ON s.id = r.supplier_id
        ORDER BY p.kind, p.name, s.name
    """)
    products = db_all("SELECT * FROM products WHERE active=1 ORDER BY kind, name")
    suppliers = db_all("SELECT * FROM suppliers WHERE active=1 ORDER BY name")
    return render_template("admin_rules.html", rules=rules, products=products, suppliers=suppliers)

@app.route("/admin/rules/create", methods=["POST"])
def admin_rules_create():
    ret = require_role("admin")
    if ret: return ret
    product_id = request.form.get("product_id", type=int)
    supplier_id = request.form.get("supplier_id", type=int)
    max_price = request.form.get("max_price", type=float)
    if not product_id or not supplier_id or max_price is None:
        flash("Dados inválidos.", "error"); return redirect(url_for("admin_rules"))
    try:
        db_exec("INSERT INTO rules (product_id, supplier_id, max_price, active) VALUES (:p,:s,:m,1)",
                p=product_id, s=supplier_id, m=max_price)
        audit("rule_create", f"product={product_id} supplier={supplier_id} max={max_price}"); flash("Regra criada.", "success")
    except Exception:
        flash("Essa combinação já existe.", "error")
    return redirect(url_for("admin_rules"))

@app.route("/admin/rules/<int:rid>/toggle", methods=["POST"])
def admin_rules_toggle(rid):
    ret = require_role("admin")
    if ret: return ret
    r = db_one("SELECT * FROM rules WHERE id=:id", id=rid)
    if not r: flash("Regra não encontrada.", "error"); return redirect(url_for("admin_rules"))
    new_active = 0 if r["active"] else 1
    db_exec("UPDATE rules SET active=:a WHERE id=:id", a=new_active, id=rid)
    audit("rule_toggle", f"id={rid} active={new_active}")
    return redirect(url_for("admin_rules"))

@app.route("/admin/rules/<int:rid>/delete", methods=["POST"])
def admin_rules_delete(rid):
    ret = require_role("admin")
    if ret: return ret
    try:
        db_exec("DELETE FROM rules WHERE id=:id", id=rid)
        audit("rule_delete", f"id={rid}")
        flash("Regra excluída.", "success")
    except Exception as e:
        flash(f"Falha ao excluir regra: {e}", "error")
    return redirect(url_for("admin_rules"))

# -------- Importação em massa (ADMIN) --------

@app.route("/admin/import/template.xlsx")
def admin_import_template():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        html = """
        {% extends "base.html" %}
        {% block title %}Template de Importação{% endblock %}
        {% block content %}
        <div class="container" style="max-width:800px;margin:0 auto">
          <h2>Template de Importação</h2>
          <p style="color:#b00"><strong>Dependência ausente:</strong> o servidor não tem <code>openpyxl</code> instalado.</p>
          <p>Adicione <code>openpyxl</code> ao seu <code>requirements.txt</code> e faça o deploy novamente:</p>
          <pre>openpyxl==3.1.5</pre>
        </div>
        {% endblock %}
        """
        return render_template_string(html)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Suppliers"
    ws1.append(["name", "active", "billing"])
    ws1.append(["Fornecedor Exemplo A", 1, 1])
    ws1.append(["Fornecedor Exemplo B", 1, 0])
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    ws2 = wb.create_sheet("Products")
    ws2.append(["name", "code", "kind", "active", "in_stock"])
    ws2.append(["Lente Asférica 1.67", "LA167", "lente", 1, 0])
    ws2.append(["Bloco Base 4", "BB4", "bloco", 1, 1])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    ws3 = wb.create_sheet("Rules")
    ws3.append(["product_name", "product_kind", "supplier_name", "max_price", "active"])
    ws3.append(["Lente Asférica 1.67", "lente", "Fornecedor Exemplo A", 250.00, 1])
    ws3.append(["Bloco Base 4", "bloco", "Fornecedor Exemplo B", 80.00, 1])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="optec_import_template.xlsx")

@app.route("/admin/import", methods=["GET", "POST"])
def admin_import():
    ret = require_role("admin")
    if ret: return ret

    report = {"suppliers": {"inserted":0, "updated":0},
              "products": {"inserted":0, "updated":0},
              "rules": {"inserted":0, "updated":0},
              "errors": []}

    if request.method == "POST":
        file = request.files.get("file")
        if not file or file.filename == "":
            flash("Envie um arquivo .xlsx", "error")
        else:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file, data_only=True)
                with engine.begin() as conn:
                    # Suppliers
                    if "Suppliers" in wb.sheetnames:
                        ws = wb["Suppliers"]
                        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
                        def idx(col): return headers.index(col) if col in headers else -1
                        i_name = idx("name"); i_active = idx("active"); i_billing = idx("billing")
                        if i_name == -1:
                            report["errors"].append("Suppliers: coluna obrigatória 'name' não encontrada.")
                        else:
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if row is None: continue
                                name = (row[i_name] or "").strip() if row[i_name] else ""
                                if not name: continue
                                active = int(row[i_active]) if (i_active != -1 and row[i_active] is not None) else 1
                                billing = int(row[i_billing]) if (i_billing != -1 and row[i_billing] is not None) else 0
                                res = conn.execute(text("""
                                    INSERT INTO suppliers (name, active, billing)
                                    VALUES (:n, :a, :b)
                                    ON CONFLICT (name) DO UPDATE SET active=EXCLUDED.active, billing=EXCLUDED.billing
                                    RETURNING (xmax = 0) AS inserted
                                """), dict(n=name, a=active, b=billing))
                                inserted = res.fetchone()[0]
                                if inserted: report["suppliers"]["inserted"] += 1
                                else: report["suppliers"]["updated"] += 1

                    # Products
                    if "Products" in wb.sheetnames:
                        ws = wb["Products"]
                        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
                        def idx(col): return headers.index(col) if col in headers else -1
                        i_name = idx("name"); i_code = idx("code"); i_kind = idx("kind"); i_active = idx("active"); i_stock = idx("in_stock")
                        if i_name == -1 or i_kind == -1:
                            report["errors"].append("Products: colunas obrigatórias 'name' e 'kind' não encontradas.")
                        else:
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if row is None: continue
                                name = (row[i_name] or "").strip() if row[i_name] else ""
                                if not name: continue
                                code = (row[i_code] or "").strip() if (i_code != -1 and row[i_code]) else ""
                                kind = (row[i_kind] or "").strip().lower() if row[i_kind] else ""
                                if kind not in ("lente", "bloco"):
                                    report["errors"].append(f"Products: kind inválido '{kind}' para '{name}'. Use 'lente' ou 'bloco'.")
                                    continue
                                active = int(row[i_active]) if (i_active != -1 and row[i_active] is not None) else 1
                                in_stock = int(row[i_stock]) if (i_stock != -1 and row[i_stock] is not None) else 0
                                res = conn.execute(text("""
                                    INSERT INTO products (name, code, kind, active, in_stock)
                                    VALUES (:n, :c, :k, :a, :instock)
                                    ON CONFLICT (name, kind) DO UPDATE SET code=EXCLUDED.code, active=EXCLUDED.active, in_stock=EXCLUDED.in_stock
                                    RETURNING (xmax = 0) AS inserted
                                """), dict(n=name, c=code, k=kind, a=active, instock=in_stock))
                                inserted = res.fetchone()[0]
                                if inserted: report["products"]["inserted"] += 1
                                else: report["products"]["updated"] += 1

                    # Rules
                    if "Rules" in wb.sheetnames:
                        ws = wb["Rules"]
                        headers = [str(c.value).strip().lower() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))]
                        def idx(col): return headers.index(col) if col in headers else -1
                        i_pn = idx("product_name"); i_pk = idx("product_kind"); i_sn = idx("supplier_name"); i_mp = idx("max_price"); i_active = idx("active")
                        if i_pn == -1 or i_pk == -1 or i_sn == -1 or i_mp == -1:
                            report["errors"].append("Rules: colunas obrigatórias 'product_name', 'product_kind', 'supplier_name', 'max_price' não encontradas.")
                        else:
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                if row is None: continue
                                pn = (row[i_pn] or "").strip() if row[i_pn] else ""
                                pk = (row[i_pk] or "").strip().lower() if row[i_pk] else ""
                                sn = (row[i_sn] or "").strip() if row[i_sn] else ""
                                try:
                                    mp = float(row[i_mp]) if row[i_mp] is not None else None
                                except:
                                    mp = None
                                if not pn or pk not in ("lente","bloco") or not sn or mp is None:
                                    report["errors"].append(f"Rules: dados inválidos (produto='{pn}', kind='{pk}', fornecedor='{sn}', max_price='{row[i_mp]}').")
                                    continue
                                active = int(row[i_active]) if (i_active != -1 and row[i_active] is not None) else 1

                                # Garantir IDs
                                prod = conn.execute(text("SELECT id FROM products WHERE name=:n AND kind=:k"), dict(n=pn, k=pk)).mappings().first()
                                if not prod:
                                    prod = conn.execute(text("""
                                        INSERT INTO products (name, code, kind, active)
                                        VALUES (:n, '', :k, 1)
                                        ON CONFLICT (name, kind) DO NOTHING
                                        RETURNING id
                                    """), dict(n=pn, k=pk)).mappings().first()
                                    if not prod:
                                        prod = conn.execute(text("SELECT id FROM products WHERE name=:n AND kind=:k"), dict(n=pn, k=pk)).mappings().first()
                                supp = conn.execute(text("SELECT id FROM suppliers WHERE name=:n"), dict(n=sn)).mappings().first()
                                if not supp:
                                    supp = conn.execute(text("""
                                        INSERT INTO suppliers (name, active)
                                        VALUES (:n, 1)
                                        ON CONFLICT (name) DO NOTHING
                                        RETURNING id
                                    """), dict(n=sn)).mappings().first()
                                    if not supp:
                                        supp = conn.execute(text("SELECT id FROM suppliers WHERE name=:n"), dict(n=sn)).mappings().first()

                                if not prod or not supp:
                                    report["errors"].append(f"Rules: não foi possível identificar produto/fornecedor ('{pn}'/'{pk}' | '{sn}').")
                                    continue

                                res = conn.execute(text("""
                                    INSERT INTO rules (product_id, supplier_id, max_price, active)
                                    VALUES (:p, :s, :m, :a)
                                    ON CONFLICT (product_id, supplier_id) DO UPDATE SET max_price=EXCLUDED.max_price, active=EXCLUDED.active
                                    RETURNING (xmax = 0) AS inserted
                                """), dict(p=prod["id"], s=supp["id"], m=mp, a=active))
                                inserted = res.fetchone()[0]
                                if inserted: report["rules"]["inserted"] += 1
                                else: report["rules"]["updated"] += 1

                flash("Importação concluída.", "success")
            except ImportError:
                report["errors"].append("Dependência ausente: instale 'openpyxl' no servidor.")
                flash("Instale 'openpyxl' para importar planilhas .xlsx.", "error")
            except Exception as e:
                report["errors"].append(str(e))
                flash("Falha na importação. Veja os erros.", "error")

    html = """
    {% extends "base.html" %}
    {% block title %}Importação em Massa
      <hr style="margin:20px 0" />

      <!-- NOVA SEÇÃO: REGRAS DO ORÇAMENTO -->
      <h3>Importar Regras do Orçamento</h3>
      <form method="POST" enctype="multipart/form-data" action="{{ url_for('admin_import_orcamento') }}">
        <div style="display:flex; gap:12px; align-items:center; flex-wrap:wrap;">
          <input type="file" name="file_orcamento" accept=".xlsx,.xls" required />
          <button class="btn" type="submit">Importar regras</button>
        </div>
      </form>

      {% if imp_orcamento %}
      <div class="card" style="margin-top:12px;">
        <div><strong>Produtos</strong>: inseridos {{ imp_orcamento.prod_inserted }}, atualizados {{ imp_orcamento.prod_updated }}</div>
        <div><strong>Serviços obrigatórios</strong>: {{ imp_orcamento.serv_obrig_upserts }}</div>
        <div><strong>Serviços opcionais</strong>: {{ imp_orcamento.serv_opc_upserts }}</div>
        <div><strong>Acréscimos</strong>: {{ imp_orcamento.acresc_upserts }}</div>
        <div class="muted" style="margin-top:6px;">Linhas lidas: {{ imp_orcamento.rows }}</div>
      </div>
      {% endif %}

      <p class="muted" style="margin-top:8px;">
        Cabeçalhos esperados (1ª aba): <em>Produto</em>, <em>Código</em>, <em>Valor</em>, <em>Tipo de visão</em> (VS/Progressiva/Bifocal),
        <em>Antirreflexo</em>, <em>Fotosensível</em>, <em>Filtro azul</em>, <em>ESF mínimo</em>, <em>ESF máximo</em>,
        <em>CIL mínimo</em>, <em>CIL máximo</em>, <em>Serviços obrigatórios</em>, <em>Serviços disponíveis</em>, <em>Acréscimos</em>.
      </p>

    {% endblock %}
    {% block content %}
    <div class="container" style="max-width: 900px; margin: 0 auto;">
      <div style="display:flex;justify-content:space-between;align-items:center;gap:16px;">
        <h2>Importar planilha (Excel .xlsx)</h2>
        <a class="btn btn-sm btn-primary" href="{{ url_for('admin_import_template') }}">Baixar Template</a>
      </div>
      <p>Use o modelo com abas <strong>Suppliers</strong> (com <code>billing</code>), <strong>Products</strong> e <strong>Rules</strong>.</p>
      <form method="post" enctype="multipart/form-data" style="margin-top: 16px;">
        <input type="file" name="file" accept=".xlsx" required />
        <button type="submit">Importar</button>
      </form>
      {% if report %}
      <hr/>
      <h3>Resultado</h3>
      <ul>
        <li>Fornecedores: {{ report.suppliers.inserted }} inseridos, {{ report.suppliers.updated }} atualizados</li>
        <li>Produtos: {{ report.products.inserted }} inseridos, {{ report.products.updated }} atualizados</li>
        <li>Regras: {{ report.rules.inserted }} inseridos, {{ report.rules.updated }} atualizados</li>
      </ul>
      {% if report.errors and report.errors|length > 0 %}
        <h4>Erros</h4>
        <ul>
          {% for e in report.errors %}
            <li style="color:#b00">{{ e }}</li>
          {% endfor %}
        </ul>
      {% endif %}
      {% endif %}
    </div>
    
      <hr style="margin:20px 0" />

      <!-- NOVA SEÇÃO: REGRAS DO ORÇAMENTO -->
      <h3>Importar Regras do Orçamento</h3>
      <form method="POST" enctype="multipart/form-data" action="{{ url_for('admin_import_orcamento') }}">
        <div style="display:flex; gap:12px; align-items:center; flex-wrap:wrap;">
          <input type="file" name="file_orcamento" accept=".xlsx,.xls" required />
          <button class="btn" type="submit">Importar regras</button>
        </div>
      </form>

      {% if imp_orcamento %}
      <div class="card" style="margin-top:12px;">
        <div><strong>Produtos</strong>: inseridos {{ imp_orcamento.prod_inserted }}, atualizados {{ imp_orcamento.prod_updated }}</div>
        <div><strong>Serviços obrigatórios</strong>: {{ imp_orcamento.serv_obrig_upserts }}</div>
        <div><strong>Serviços opcionais</strong>: {{ imp_orcamento.serv_opc_upserts }}</div>
        <div><strong>Acréscimos</strong>: {{ imp_orcamento.acresc_upserts }}</div>
        <div class="muted" style="margin-top:6px;">Linhas lidas: {{ imp_orcamento.rows }}</div>
      </div>
      {% endif %}

      <p class="muted" style="margin-top:8px;">
        Cabeçalhos esperados (1ª aba): <em>Produto</em>, <em>Código</em>, <em>Valor</em>, <em>Tipo de visão</em> (VS/Progressiva/Bifocal),
        <em>Antirreflexo</em>, <em>Fotosensível</em>, <em>Filtro azul</em>, <em>ESF mínimo</em>, <em>ESF máximo</em>,
        <em>CIL mínimo</em>, <em>CIL máximo</em>, <em>Serviços obrigatórios</em>, <em>Serviços disponíveis</em>, <em>Acréscimos</em>.
      </p>

    {% endblock %}
    """
    imp_data = session.pop('imp_orcamento', None)
    return render_template_string(html, report=report, imp_orcamento=imp_data)

# -------- Comprador: Novo Pedido (com troca automática por cilindro) --------


@app.post("/admin/import_orcamento")
def admin_import_orcamento():
    ret = require_role("admin")
    if ret: return ret

    from flask import redirect, url_for
    try:
        from openpyxl import load_workbook
    except Exception:
        flash("Dependência ausente: instale 'openpyxl'.", "error")
        return redirect(url_for('admin_import'))

    file = request.files.get("file_orcamento")
    if not file or file.filename == "":
        flash("Selecione um arquivo .xlsx", "error")
        return redirect(url_for('admin_import'))

    # helpers
    import re
    from decimal import Decimal, InvalidOperation
    def _to_dec(v, default=0):
        if v is None: return Decimal(default)
        if isinstance(v, (int,float,Decimal)): return Decimal(str(v))
        s = str(v).strip().replace(",", ".")
        try: return Decimal(s)
        except Exception: return Decimal(default)
    def _b01(v):
        if v is None: return 0
        s = str(v).strip().lower()
        return 1 if s in ("1","sim","s","true","t","x") else 0
    def _split_codes(s):
        if not s: return []
        parts = re.split(r"[.;,]\s*", str(s).strip())
        return [p.strip() for p in parts if p.strip()]
    def _parse_acresc(expr):
        if not expr: return []
        out = []
        for it in re.split(r"[.;,]\s*", str(expr).strip()):
            if not it: continue
            m = re.match(r"^\s*([A-Za-z0-9]+)\s*(?:\[(.*?)\])?\s*$", it)
            if not m: 
                out.append((it, None, None, None, None)); continue
            code, body = m.group(1), m.group(2)
            esf_min = esf_max = cil_min = cil_max = None
            if body:
                for part in re.split(r"\s*;\s*", body):
                    pm = re.match(r"^(esf|cil)\s*:\s*([+\-]?\d+(?:[.,]\d+)?)\s*(?:\.\.|a|to|-)\s*([+\-]?\d+(?:[.,]\d+)?)\s*$", part, flags=re.I)
                    if pm:
                        kind = pm.group(1).lower()
                        v1 = _to_dec(pm.group(2)); v2 = _to_dec(pm.group(3))
                        lo, hi = (v1 if v1<=v2 else v2), (v2 if v2>=v1 else v1)
                        if kind == "esf": esf_min, esf_max = lo, hi
                        else: cil_min, cil_max = lo, hi
            out.append((code, esf_min, esf_max, cil_min, cil_max))
        return out

    try:
        wb = load_workbook(filename=file, data_only=True)
        ws = wb.worksheets[0]
    except Exception as e:
        flash(f"Erro ao ler Excel: {e}", "error")
        return redirect(url_for('admin_import'))

    # header map
    header_map = {}
    for j, c in enumerate(ws[1], start=1):
        header_map[(str(c.value or '').strip().lower())] = j
    def col(*names):
        for n in names:
            n=n.lower()
            if n in header_map: return header_map[n]
        return None

    idx_prod   = col("produto","nome")
    idx_code   = col("código","codigo","cod")
    idx_valor  = col("valor","preço","preco")
    idx_visao  = col("tipo de visão","tipo de visao","visao","visão","tv")
    idx_ar     = col("antirreflexo","anti-reflexo","ar")
    idx_foto   = col("fotosensível","fotossensivel","foto")
    idx_azul   = col("filtro azul","azul","blue")
    idx_esfmin = col("esf mínimo","esf minimo","esf min","esf_min")
    idx_esfmax = col("esf máximo","esf maximo","esf max","esf_max")
    idx_cilmin = col("cil mínimo","cil minimo","cil min","cil_min")
    idx_cilmax = col("cil máximo","cil maximo","cil max","cil_max")
    idx_servob = col("serviços obrigatórios","servicos obrigatorios","obrigatórios","obrigatorios","n1-14")
    idx_servop = col("serviços disponíveis","servicos disponiveis","disponíveis","disponiveis","o15")
    idx_acresc = col("acréscimos","acrescimos","acréscimo","acrescimo")

    must = [idx_prod,idx_code,idx_valor,idx_visao,idx_ar,idx_foto,idx_azul,idx_esfmin,idx_esfmax,idx_cilmin,idx_cilmax,idx_servob,idx_servop]
    if not all(must):
        flash("Cabeçalhos ausentes na 1ª aba. Confira os nomes.", "error")
        return redirect(url_for('admin_import'))

    prod_ins = prod_upd = 0
    serv_ob_upserts = serv_opc_upserts = acresc_upserts = 0
    rows = 0

    with engine.begin() as conn:
        for i in range(2, ws.max_row + 1):
            rows += 1
            nome = (ws.cell(i, idx_prod).value or "").strip()
            code = (str(ws.cell(i, idx_code).value or "").strip())
            if not nome or not code: 
                continue

            vis_raw = (str(ws.cell(i, idx_visao).value or "").strip().lower())
            if vis_raw in ("vs","visao simples","visão simples","visao_simples","visão_simples"):
                visao = "visao_simples"
            elif vis_raw in ("progressiva",): visao = "progressiva"
            elif vis_raw in ("bifocal",):     visao = "bifocal"
            else: visao = "visao_simples"

            price  = _to_dec(ws.cell(i, idx_valor).value, 0)
            ar     = _b01(ws.cell(i, idx_ar).value)
            foto   = _b01(ws.cell(i, idx_foto).value)
            azul   = _b01(ws.cell(i, idx_azul).value)
            esfmin = _to_dec(ws.cell(i, idx_esfmin).value, 0)
            esfmax = _to_dec(ws.cell(i, idx_esfmax).value, 0)
            cilmin = _to_dec(ws.cell(i, idx_cilmin).value, 0)
            cilmax = _to_dec(ws.cell(i, idx_cilmax).value, 0)

            r = conn.execute(text("""
              INSERT INTO orc_produto (code,name,price,visao,ar,foto,azul,esf_min,esf_max,cil_min,cil_max,updated_at)
              VALUES (:code,:name,:price,:visao,:ar,:foto,:azul,:esfmin,:esfmax,:cilmin,:cilmax,NOW())
              ON CONFLICT (code) DO UPDATE SET
                name=EXCLUDED.name, price=EXCLUDED.price, visao=EXCLUDED.visao,
                ar=EXCLUDED.ar, foto=EXCLUDED.foto, azul=EXCLUDED.azul,
                esf_min=EXCLUDED.esf_min, esf_max=EXCLUDED.esf_max,
                cil_min=EXCLUDED.cil_min, cil_max=EXCLUDED.cil_max,
                updated_at=NOW()
              RETURNING (xmax = 0) AS inserted, id
            """), {"code":code,"name":nome,"price":price,"visao":visao,"ar":ar,"foto":foto,"azul":azul,
                   "esfmin":esfmin,"esfmax":esfmax,"cilmin":cilmin,"cilmax":cilmax})
            row = r.fetchone()
            inserted = bool(row[0]); pid = row[1]
            prod_ins += 1 if inserted else 0
            prod_upd += 0 if inserted else 1

            for sc in _split_codes(ws.cell(i, idx_servob).value):
                conn.execute(text("INSERT INTO orc_servico_catalogo (code) VALUES (:c) ON CONFLICT DO NOTHING"), {"c": sc})
                conn.execute(text("""
                  INSERT INTO orc_produto_serv_obrig (produto_id, serv_code)
                  VALUES (:pid, :sc)
                  ON CONFLICT (produto_id, serv_code) DO NOTHING
                """), {"pid": pid, "sc": sc})
                serv_ob_upserts += 1

            for sc in _split_codes(ws.cell(i, idx_servop).value):
                conn.execute(text("INSERT INTO orc_servico_catalogo (code) VALUES (:c) ON CONFLICT DO NOTHING"), {"c": sc})
                conn.execute(text("""
                  INSERT INTO orc_produto_serv_opc (produto_id, serv_code)
                  VALUES (:pid, :sc)
                  ON CONFLICT (produto_id, serv_code) DO NOTHING
                """), {"pid": pid, "sc": sc})
                serv_opc_upserts += 1

            if idx_acresc:
                for code_ac, esf_min, esf_max, cil_min, cil_max in _parse_acresc(ws.cell(i, idx_acresc).value):
                    conn.execute(text("INSERT INTO orc_servico_catalogo (code) VALUES (:c) ON CONFLICT DO NOTHING"),
                                 {"c": code_ac})
                    conn.execute(text("""
                      INSERT INTO orc_produto_acrescimo (produto_id, serv_code, esf_min, esf_max, cil_min, cil_max)
                      VALUES (:pid,:sc,:esfmin,:esfmax,:cilmin,:cilmax)
                      ON CONFLICT (produto_id, serv_code, esf_min, esf_max, cil_min, cil_max) DO NOTHING
                    """), {"pid":pid,"sc":code_ac,
                             "esfmin":esf_min,"esfmax":esf_max,"cilmin":cil_min,"cilmax":cil_max})
                    acresc_upserts += 1

    session['imp_orcamento'] = {
        "prod_inserted": prod_ins, "prod_updated": prod_upd,
        "serv_obrig_upserts": serv_ob_upserts, "serv_opc_upserts": serv_opc_upserts,
        "acresc_upserts": acresc_upserts, "rows": rows
    }
    flash("Importação das regras do orçamento concluída.", "success")
    return redirect(url_for('admin_import'))

@app.route("/compras/novo", methods=["GET","POST"])
def compras_novo():
    ret = require_role("comprador","admin")
    if ret: return ret

    combos = db_all("""
        SELECT r.id as rule_id, p.id as product_id, p.name as product_name, p.code as product_code, p.kind,
               s.id as supplier_id, s.name as supplier_name, r.max_price
        FROM rules r
        JOIN products p ON p.id = r.product_id
        JOIN suppliers s ON s.id = r.supplier_id
        WHERE r.active=1 AND p.active=1 AND s.active=1
        ORDER BY s.name, p.kind, p.name
    """)
    products = db_all("SELECT id, name, code, kind FROM products WHERE active=1 ORDER BY kind, name")

    combos = [dict(r) for r in combos]
    products = [dict(p) for p in products]

    if request.method == "POST":

        # === NOVO (lote): se vier uma lista JSON completa no campo oculto, processa todos os itens ===
        raw_payload = (request.form.get("hidden_payload") or "").strip()
        if raw_payload:
            try:
                import json
                from collections import defaultdict, Counter
                from sqlalchemy import text as _t
                payload = json.loads(raw_payload)
                bulk_items = []
                for it in (payload or []):
                    try:
                        bulk_items.append({
                            "os_number": str(it.get("os_number","")).strip(),
                            "product_id": int(it.get("product_id")),
                            "supplier_id": int(it.get("supplier_id")),
                            "price": float(it.get("price")),
                            "tipo": str(it.get("tipo","")).lower(),  # 'lente' ou 'bloco'
                            "d": {
                                "sphere": it.get("sphere"),
                                "cylinder": (None if it.get("cylinder") is None else -abs(float(it.get("cylinder")))),
                                "base": it.get("base"),
                                "addition": it.get("addition"),
                            }
                        })
                    except Exception:
                        pass
                # validações básicas
                def _step_ok(x: float) -> bool:
                    try:
                        return (abs(float(x) * 100) % 25) == 0
                    except Exception:
                        return False
                def _validate_item(pid, sid, tipo_item, price, d):
                    rule = db_one("""
                        SELECT r.*, p.kind as product_kind
                        FROM rules r JOIN products p ON p.id = r.product_id
                        WHERE r.product_id=:pid AND r.supplier_id=:sid AND r.active=1
                    """, pid=pid, sid=sid)
                    if not rule:
                        return None, None, "Fornecedor indisponível para este produto."
                    if price is None or price <= 0 or price > float(rule["max_price"]) + 1e-6:
                        return None, None, f"Preço inválido ou acima do máximo (R$ {float(rule['max_price']):.2f})."
                    s = d.get("sphere"); c = d.get("cylinder")
                    if s is None or float(s) < -20 or float(s) > 20 or (not _step_ok(s)):
                        return None, None, "Esférico inválido (−20 a +20 em passos de 0,25)."
                    if c is None or float(c) > 0 or float(c) < -15 or (not _step_ok(c)):
                        return None, None, "Cilíndrico inválido (0 até −15 em passos de 0,25)."
                    if tipo_item == "lente":
                        new_pid, new_price, _changed = maybe_swap_lente_by_cylinder(pid, sid, price, c)
                        return new_pid, new_price, None
                    else:
                        return pid, price, None
                if bulk_items:
                    # limite 2 por OS ( somando o que já existe )
                    from collections import Counter
                    os_new = Counter([it["os_number"] for it in bulk_items if it.get("os_number")])
                    for osn, add_n in os_new.items():
                        row = db_one("SELECT COUNT(*) AS n FROM purchase_items WHERE os_number=:os", os=osn)
                        existing_n = int(row["n"] if row else 0)
                        if existing_n + add_n > 2:
                            flash(f"OS {osn} excede o limite de 2 unidades.", "error")
                            return render_template("compras_novo.html", combos=combos, products=products)
                    validated = []
                    for it in bulk_items:
                        if not it["os_number"]:
                            flash("Há item sem número de OS.", "error")
                            return render_template("compras_novo.html", combos=combos, products=products)
                        pid, price_adj, err = _validate_item(it["product_id"], it["supplier_id"], it["tipo"], it["price"], it["d"])
                        if err:
                            flash(f"OS {it['os_number']}: {err}", "error")
                            return render_template("compras_novo.html", combos=combos, products=products)
                        validated.append({
                            "product_id": pid, "supplier_id": it["supplier_id"], "price": price_adj,
                            "d": it["d"], "os_number": it["os_number"]
                        })
                    # cria 1 pedido por fornecedor
                    by_supplier = defaultdict(list)
                    for it in validated:
                        by_supplier[it["supplier_id"]].append(it)
                    created_orders = []
                    from datetime import datetime
                    with engine.begin() as conn:
                        for sup_id, its in by_supplier.items():
                            supplier_row = conn.execute(_t("SELECT * FROM suppliers WHERE id=:id"), {"id": sup_id}).mappings().first()
                            faturado = (supplier_row and (supplier_row.get("billing") or 0) == 1)
                            total_group = sum(float(i["price"]) for i in its)
                            status = 'PAGO' if faturado else 'PENDENTE_PAGAMENTO'
                            os_list = ", ".join(sorted({i["os_number"] for i in its}))
                            note = f"OS {os_list}"
                            res = conn.execute(_t("""
                                INSERT INTO purchase_orders (buyer_id, supplier_id, status, total, note, created_at, updated_at)
                                VALUES (:b,:s,:st,:t,:n,:c,:u) RETURNING id
                            """), {"b": session["user_id"], "s": sup_id, "st": status, "t": total_group, "n": note,
                                    "c": datetime.utcnow(), "u": datetime.utcnow()})
                            order_id = res.scalar_one()
                            for i in its:
                                conn.execute(_t("""
                                    INSERT INTO purchase_items (order_id, product_id, quantity, unit_price, sphere, cylinder, base, addition, os_number)
                                    VALUES (:o,:p,1,:pr,:sf,:cl,:ba,:ad,:os)
                                """), {"o": order_id, "p": i["product_id"], "pr": i["price"],
                                        "sf": i["d"].get("sphere"), "cl": i["d"].get("cylinder"),
                                        "ba": i["d"].get("base"), "ad": i["d"].get("addition"), "os": i["os_number"]})
                            if faturado:
                                conn.execute(_t("""
                                    INSERT INTO payments (order_id, payer_id, method, reference, paid_at, amount)
                                    VALUES (:o,:p,:m,:r,:d,:a)
                                """), {"o": order_id, "p": session["user_id"], "m": "FATURADO",
                                        "r": note, "d": datetime.utcnow(), "a": total_group})
                            created_orders.append((order_id, faturado))
                    for oid, fat in created_orders: audit("order_create", f"id={oid} faturado={int(fat)}")
                    if any(f for _, f in created_orders) and any((not f) for _, f in created_orders):
                        flash("Pedidos criados: 1 FATURADO (lançado) e 1 PENDENTE (enviado ao pagador).", "success")
                    elif all(f for _, f in created_orders):
                        flash("Pedido(s) criado(s) como FATURADO(s) e incluído(s) diretamente no relatório.", "success")
                    else:
                        flash("Pedido(s) criado(s) e enviado(s) ao pagador.", "success")
                    return redirect(url_for("compras_lista"))
            except Exception as _e:
                # em qualquer erro de parsing, segue o fluxo original (form simples)
                pass
        action = (request.form.get("action") or "").strip().lower()

        os_number = (request.form.get("os_number") or "").strip()
        pair_option = request.form.get("pair_option")  # 'meio' ou 'par'
        tipo = (request.form.get("tipo") or "").lower()  # 'lente' ou 'bloco'
        product_id = request.form.get("product_id", type=int)
        product_code = (request.form.get("product_code") or "").strip()
        supplier_main = request.form.get("supplier_main", type=int)
        price_main = request.form.get("price_main", type=float)

        supplier_distinto = request.form.get("supplier_distinto") == "on"
        supplier_second = request.form.get("supplier_second", type=int) if supplier_distinto else None
        price_second = request.form.get("price_second", type=float) if supplier_distinto else None

        if not os_number:
            flash("Informe o número da OS.", "error")
            return render_template("compras_novo.html", combos=combos, products=products)

        existing = db_one("SELECT COUNT(*) AS n FROM purchase_items WHERE os_number=:os", os=os_number)
        existing_n = int(existing["n"] if existing else 0)

        if pair_option not in ("meio","par"):
            flash("Selecione se é meio par ou um par.", "error")
            return render_template("compras_novo.html", combos=combos, products=products)

        if tipo not in ("lente","bloco"):
            flash("Selecione o tipo (lente/bloco).", "error")
            return render_template("compras_novo.html", combos=combos, products=products)

        if not product_id and product_code:
            p = db_one("SELECT id FROM products WHERE code=:c AND kind=:k AND active=1", c=product_code, k=tipo)
            if p:
                product_id = int(p["id"])

        if not product_id:
            flash("Selecione o produto (ou informe um código válido).", "error")
            return render_template("compras_novo.html", combos=combos, products=products)

        rule_main = db_one("""
            SELECT r.*, p.kind as product_kind
            FROM rules r JOIN products p ON p.id = r.product_id
            WHERE r.product_id=:pid AND r.supplier_id=:sid AND r.active=1
        """, pid=product_id, sid=supplier_main)
        if not rule_main:
            flash("Fornecedor principal indisponível para este produto.", "error")
            return render_template("compras_novo.html", combos=combos, products=products)
        if price_main is None or price_main <= 0 or price_main > float(rule_main["max_price"]) + 1e-6:
            flash(f"Preço do item principal inválido ou acima do máximo (R$ {float(rule_main['max_price']):.2f}).", "error")
            return render_template("compras_novo.html", combos=combos, products=products)

        def _step_ok(x: float) -> bool:
            return (abs(x * 100) % 25) == 0

        def validate_lente(prefix):
            sphere = request.form.get(f"{prefix}_sphere", type=float)
            cylinder_raw = request.form.get(f"{prefix}_cylinder", type=float)
            cylinder = None
            if cylinder_raw is not None:
                cylinder = -abs(cylinder_raw)
            if sphere is None or sphere < -20 or sphere > 20 or not _step_ok(sphere):
                return None, "Esférico inválido (−20 a +20 em passos de 0,25)."
            if cylinder is None or cylinder > 0 or cylinder < -15 or not _step_ok(cylinder):
                return None, "Cilíndrico inválido (0 até −15 em passos de 0,25)."
            return {"sphere": sphere, "cylinder": cylinder, "base": None, "addition": None}, None

        def validate_bloco(prefix):
            base = request.form.get(f"{prefix}_base", type=float)
            addition = request.form.get(f"{prefix}_addition", type=float)
            allowed_bases = {0.5,1.0,2.0,4.0,6.0,8.0,10.0}
            if base is None or base not in allowed_bases:
                return None, "Base inválida (0,5; 1; 2; 4; 6; 8; 10)."
            # Permitir 0 para BVS (bloco visão simples). Se None, tratar como 0.
            addition = 0.0 if addition is None else addition
            # Válido se: 0 (BVS) OU entre 1 e 4 (inclusive), sempre em passos de 0,25
            if addition < 0 or not _step_ok(addition) or (addition != 0 and not (1.0 <= addition <= 4.0)):
                return None, "Adição inválida: use 0 para BVS (bloco) ou +1,00 a +4,00 em passos de 0,25."
            return {"sphere": None, "cylinder": None, "base": base, "addition": addition}, None

        items_to_add = []

        if tipo == "lente":
            d1, err = validate_lente("d1")
            if err:
                flash(err, "error"); return render_template("compras_novo.html", combos=combos, products=products)
        else:
            d1, err = validate_bloco("d1")
            if err:
                flash(err, "error"); return render_template("compras_novo.html", combos=combos, products=products)

        # Troca automática por cilindro (Item A)
        pid1, price1 = product_id, price_main
        if tipo == "lente":
            pid1, price1, changed1 = maybe_swap_lente_by_cylinder(product_id, supplier_main, price_main, d1["cylinder"])
            if changed1:
                flash(f"Item (A) ajustado por cilindro: {changed1[0]} → {changed1[1]}.", "info")

        items_to_add.append({"product_id": pid1, "supplier_id": supplier_main, "price": price1, "d": d1})

        if pair_option == "par":
            if supplier_distinto:
                if not supplier_second:
                    flash("Selecione o fornecedor do segundo item.", "error"); return render_template("compras_novo.html", combos=combos, products=products)
                rule_second = db_one("""
                    SELECT r.*, p.kind as product_kind
                    FROM rules r JOIN products p ON p.id = r.product_id
                    WHERE r.product_id=:pid AND r.supplier_id=:sid AND r.active=1
                """, pid=product_id, sid=supplier_second)
                if not rule_second:
                    flash("Fornecedor do segundo item indisponível para este produto.", "error"); return render_template("compras_novo.html", combos=combos, products=products)
                if price_second is None or price_second <= 0 or price_second > float(rule_second["max_price"]) + 1e-6:
                    flash(f"Preço do segundo item inválido ou acima do máximo (R$ {float(rule_second['max_price']):.2f}).", "error"); return render_template("compras_novo.html", combos=combos, products=products)
            else:
                supplier_second, price_second = supplier_main, price_main

            if tipo == "lente":
                d2, err = validate_lente("d2")
                if err:
                    flash(err, "error"); return render_template("compras_novo.html", combos=combos, products=products)
            else:
                d2, err = validate_bloco("d2")
                if err:
                    flash(err, "error"); return render_template("compras_novo.html", combos=combos, products=products)

            # Troca automática por cilindro (Item B)
            pid2, price2 = product_id, price_second
            if tipo == "lente":
                pid2, price2, changed2 = maybe_swap_lente_by_cylinder(product_id, supplier_second, price_second, d2["cylinder"])
                if changed2:
                    flash(f"Item (B) ajustado por cilindro: {changed2[0]} → {changed2[1]}.", "info")

            items_to_add.append({"product_id": pid2, "supplier_id": supplier_second, "price": price2, "d": d2})

        if existing_n + len(items_to_add) > 2:
            flash("Cada número de OS só pode ter no máximo um par (2 unidades).", "error")
            return render_template("compras_novo.html", combos=combos, products=products)

        # === CORREÇÃO: criar 1 pedido por fornecedor ===
        from collections import defaultdict
        by_supplier = defaultdict(list)
        for it in items_to_add:
            by_supplier[it["supplier_id"]].append(it)

        created_orders = []
        with engine.begin() as conn:
            for sup_id, its in by_supplier.items():
                supplier_row = conn.execute(text("SELECT * FROM suppliers WHERE id=:id"), dict(id=sup_id)).mappings().first()
                faturado = (supplier_row and (supplier_row.get("billing") or 0) == 1)

                total_group = sum(float(it["price"]) for it in its)
                status = 'PAGO' if faturado else 'PENDENTE_PAGAMENTO'

                res = conn.execute(text("""
                    INSERT INTO purchase_orders (buyer_id, supplier_id, status, total, note, created_at, updated_at)
                    VALUES (:b,:s,:st,:t,:n,:c,:u) RETURNING id
                """), dict(b=session["user_id"], s=sup_id, st=status, t=total_group,
                           n=f"OS {os_number} ({pair_option})", c=datetime.utcnow(), u=datetime.utcnow()))
                order_id = res.scalar_one()

                for it in its:
                    conn.execute(text("""
                        INSERT INTO purchase_items (order_id, product_id, quantity, unit_price, sphere, cylinder, base, addition, os_number)
                        VALUES (:o,:p,1,:pr,:sf,:cl,:ba,:ad,:os)
                    """), dict(o=order_id, p=it["product_id"], pr=it["price"],
                               sf=it["d"]["sphere"], cl=it["d"]["cylinder"],
                               ba=it["d"]["base"], ad=it["d"]["addition"], os=os_number))

                if faturado:
                    conn.execute(text("""
                        INSERT INTO payments (order_id, payer_id, method, reference, paid_at, amount)
                        VALUES (:o,:p,:m,:r,:d,:a)
                    """), dict(o=order_id, p=session["user_id"], m="FATURADO",
                               r=f"OS {os_number}", d=datetime.utcnow(), a=total_group))

                created_orders.append((order_id, faturado))

        for oid, fat in created_orders:
            audit("order_create", f"id={oid} os={os_number} faturado={int(fat)}")

        # Se o botão foi "Adicionar à lista", limpa via redirect (PRG)
        if action in ("add","adicionar","add_to_list"):
            flash("Item(ns) adicionados à lista.", "success")
            return redirect(url_for("compras_novo"))

        if any(fat for _, fat in created_orders) and any((not fat) for _, fat in created_orders):
            flash("Pedidos criados: 1 FATURADO (lançado) e 1 PENDENTE (enviado ao pagador).", "success")
        elif all(fat for _, fat in created_orders):
            flash("Pedido(s) criado(s) como FATURADO(s) e incluído(s) diretamente no relatório.", "success")
        else:
            flash("Pedido(s) criado(s) e enviado(s) ao pagador.", "success")

        return redirect(url_for("compras_lista"))

    return render_template("compras_novo.html", combos=combos, products=products)

# -------- Comprador: lista/detalhe --------

@app.route("/compras")
def compras_lista():
    ret = require_role("comprador","admin")
    if ret: return ret
    orders = db_all("""
        SELECT
            o.*,
            s.name AS supplier_name,
            CASE
              WHEN o.status = 'PAGO' AND EXISTS (
                SELECT 1 FROM payments pay
                WHERE pay.order_id = o.id AND pay.method = 'FATURADO'
              )
              THEN 'FATURADO'
              ELSE o.status
            END AS status_exibido
        FROM purchase_orders o
        JOIN suppliers s ON s.id = o.supplier_id
        WHERE o.buyer_id=:b
        ORDER BY o.id DESC
    """, b=session["user_id"])

    return render_template("compras_lista.html", orders=orders)

@app.route("/compras/<int:oid>")
def compras_detalhe(oid):
    ret = require_role("comprador","admin")
    if ret: return ret
    order = db_one("""
        SELECT o.*, s.name as supplier_name
        FROM purchase_orders o JOIN suppliers s ON s.id = o.supplier_id
        WHERE o.id=:id
    """, id=oid)
    if not order:
        flash("Pedido não encontrado.", "error"); return redirect(url_for("compras_lista"))
    if session.get("role") != "admin" and order["buyer_id"] != session.get("user_id"):
        flash("Acesso negado ao pedido.", "error"); return redirect(url_for("compras_lista"))
    items = db_all("""
        SELECT i.*, p.name as product_name, p.kind as product_kind
        FROM purchase_items i JOIN products p ON p.id = i.product_id
        WHERE i.order_id=:id ORDER BY i.id
    """, id=oid)
    return render_template("compras_detalhe.html", order=order, items=items)

# -------- Pagador --------

@app.route("/pagamentos")
def pagamentos_lista():
    ret = require_role("pagador","admin")
    if ret: return ret
    orders = db_all("""
        SELECT o.*, u.username as buyer_name, s.name as supplier_name
        FROM purchase_orders o
        JOIN users u ON u.id = o.buyer_id
        JOIN suppliers s ON s.id = o.supplier_id
        WHERE o.status='PENDENTE_PAGAMENTO'
        ORDER BY o.created_at ASC
    """)
    return render_template("pagamentos_lista.html", orders=orders, role=session.get("role"))

@app.route("/pagamentos/<int:oid>", methods=["GET","POST"])
def pagamentos_detalhe(oid):
    ret = require_role("pagador","admin")
    if ret: return ret
    order = db_one("""
        SELECT o.*, u.username as buyer_name, s.name as supplier_name
        FROM purchase_orders o
        JOIN users u ON u.id = o.buyer_id
        JOIN suppliers s ON s.id = o.supplier_id
        WHERE o.id=:id
    """, id=oid)
    items = db_all("""
        SELECT i.*, p.name as product_name, p.kind as product_kind
        FROM purchase_items i JOIN products p ON p.id = i.product_id
        WHERE i.order_id=:id
    """, id=oid)
    if not order:
        flash("Pedido não encontrado.", "error"); return redirect(url_for("pagamentos_lista"))
    if request.method == "POST":
        method = (request.form.get("method") or "PIX").strip()
        reference = (request.form.get("reference") or "").strip()
        amount = request.form.get("amount", type=float)
        if amount is None or amount <= 0:
            flash("Valor inválido.", "error"); return render_template("pagamentos_detalhe.html", order=order, items=items)
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO payments (order_id, payer_id, method, reference, paid_at, amount)
                VALUES (:o,:p,:m,:r,:d,:a)
            """), dict(o=oid, p=session["user_id"], m=method, r=reference, d=datetime.utcnow(), a=amount))
            conn.execute(text("UPDATE purchase_orders SET status='PAGO', updated_at=:u WHERE id=:id"),
                         dict(u=datetime.utcnow(), id=oid))
        audit("order_paid", f"id={oid} amount={amount}")
        flash("Pagamento registrado e pedido baixado como PAGO.", "success"); return redirect(url_for("pagamentos_lista"))
    return render_template("pagamentos_detalhe.html", order=order, items=items)

# -------- Relatórios --------

@app.route("/relatorios")
def relatorios_index():
    ret = require_role("admin","pagador")
    if ret: return ret
    hoje = date.today().isoformat()
    html = """
    {% extends "base.html" %}
    {% block title %}Relatórios{% endblock %}
    {% block content %}
    <div class="container" style="max-width: 760px; margin: 0 auto;">
      <h2>Relatórios</h2>

      <div class="card" style="padding:12px; margin-bottom:16px;">
        <h3 style="margin:0 0 8px;">Relatório Diário</h3>
        <form method="get" action="{{ url_for('relatorio_diario_xlsx') }}" style="display:flex; gap:12px; align-items:flex-end; flex-wrap:wrap;">
          <div>
            <label for="date"><strong>Data do relatório</strong></label><br/>
            <input type="date" id="date" name="date" value="{{ hoje }}"/>
          </div>
          <div>
            <button class="btn primary" type="submit">Baixar Excel (.xlsx)</button>
          </div>
          <div>
            <a class="btn" href="{{ url_for('relatorio_diario_csv', date=hoje) }}">Baixar CSV</a>
          </div>
        </form>
        <small class="muted">O Excel contém: <b>Fornecedor, Produto, Estoque, Dioptria, Data, Método, Valor</b> e o <b>TOTAL</b>.</small>
      </div>

      <div class="card" style="padding:12px;">
        <h3 style="margin:0 0 8px;">Relatório por Período</h3>
        <form method="get" action="{{ url_for('relatorio_periodo_xlsx') }}" style="display:flex; gap:12px; align-items:flex-end; flex-wrap:wrap;">
          <div>
            <label for="start"><strong>De</strong></label><br/>
            <input type="date" id="start" name="start" value="{{ hoje }}"/>
          </div>
          <div>
            <label for="end"><strong>Até</strong></label><br/>
            <input type="date" id="end" name="end" value="{{ hoje }}"/>
          </div>
          <div>
            <button class="btn primary" type="submit">Baixar Excel (.xlsx)</button>
          </div>
        </form>
        <small class="muted">Inclui pagamentos feitos e faturados (<b>FATURADO</b>) dentro do intervalo.</small>
      </div>
    </div>
    {% endblock %}
    """
    return render_template_string(html, hoje=hoje)

@app.route("/relatorios/diario.xlsx")
def relatorio_diario_xlsx():
    ret = require_role("admin","pagador")
    if ret: return ret
    day = request.args.get("date") or date.today().isoformat()
    try:
        xbytes = build_excel_bytes_for_day(day)
        return send_file(io.BytesIO(xbytes),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"pagamentos_{day}.xlsx")
    except Exception as e:
        print(f"[RELATORIO] Falha ao gerar XLSX: {e}", flush=True)
        flash("Excel indisponível no momento. Baixando em CSV.", "warning")
        return redirect(url_for("relatorio_diario_csv", date=day))

@app.route("/relatorios/diario.csv")
def relatorio_diario_csv():
    ret = require_role("admin","pagador")
    if ret: return ret
    day = request.args.get("date") or date.today().isoformat()
    rows = db_all("""
        SELECT pay.paid_at, pay.amount, pay.method, pay.reference,
               o.id as order_id, s.name as supplier_name, u.username as payer_name
        FROM payments pay
        JOIN purchase_orders o ON o.id = pay.order_id
        JOIN suppliers s ON s.id = o.supplier_id
        JOIN users u ON u.id = pay.payer_id
        WHERE DATE(pay.paid_at)=:day
        ORDER BY pay.paid_at ASC
    """, day=day)
    output = io.StringIO(); writer = csv.writer(output, lineterminator="\n")
    writer.writerow(["paid_at","amount","method","reference","order_id","supplier","payer"])
    for r in rows:
        paid_at = r["paid_at"].isoformat(sep=" ", timespec="seconds") if hasattr(r["paid_at"], "isoformat") else str(r["paid_at"])
        writer.writerow([paid_at, f"{float(r['amount']):.2f}", r["method"], r["reference"], r["order_id"], r["supplier_name"], r["payer_name"]])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode("utf-8-sig")), mimetype="text/csv; charset=utf-8",
                     as_attachment=True, download_name=f"pagamentos_{day}.csv")

@app.route("/relatorios/periodo.xlsx")
def relatorio_periodo_xlsx():
    ret = require_role("admin","pagador")
    if ret: return ret
    start = request.args.get("start") or date.today().isoformat()
    end   = request.args.get("end")   or start
    try:
        xbytes = build_excel_bytes_for_period(start, end)
        return send_file(io.BytesIO(xbytes),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"pagamentos_{start}_a_{end}.xlsx")
    except Exception as e:
        print(f"[RELATORIO-PERIODO] Falha ao gerar XLSX: {e}", flush=True)
        flash("Excel indisponível no momento para o período.", "warning")
        return redirect(url_for("relatorio_diario_csv", date=start))

# -------- Admin: excluir pedidos --------

@app.route("/admin/orders/<int:oid>/delete", methods=["POST"])
def admin_orders_delete(oid):
    ret = require_role("admin")
    if ret: return ret
    with engine.begin() as conn:
        conn.execute(text("DELETE FROM payments WHERE order_id=:id"), dict(id=oid))
        conn.execute(text("DELETE FROM purchase_items WHERE order_id=:id"), dict(id=oid))
        conn.execute(text("DELETE FROM purchase_orders WHERE id=:id"), dict(id=oid))
    audit("order_delete", f"id={oid}")
    flash("Pedido excluído.", "success")
    return redirect(url_for("compras_lista"))

# ============================ BOOTSTRAP ============================

try:
    init_db()
except Exception as e:
    print(f"[BOOT] init_db() falhou: {e}", flush=True)

if __name__ == "__main__":
    # Para rodar local, defina DATABASE_URL (ex.: sqlite:///local.db) antes de executar
    app.run(host="0.0.0.0", port=5000, debug=True)

# === Extornos ===

@app.route("/extornos", methods=["GET"])
def extornos_index():
    ret = require_role("pagador","admin")
    if ret: return ret

    day = (request.args.get("date") or date.today().isoformat())

    items = db_all("""
        SELECT
          i.id AS item_id,
          o.id AS order_id,
          s.id AS supplier_id,
          s.name AS supplier_name,
          p.name AS product_name,
          p.code AS product_code,
          i.quantity, 
          i.unit_price,
          i.sphere, i.cylinder, i.base, i.addition,
          EXISTS(SELECT 1 FROM supplier_credits c WHERE c.item_id = i.id) AS already
        FROM payments pay
        JOIN purchase_orders o ON o.id = pay.order_id
        JOIN suppliers s       ON s.id = o.supplier_id
        JOIN purchase_items i  ON i.order_id = o.id
        JOIN products p        ON p.id = i.product_id
        WHERE DATE(pay.paid_at) = :day
        ORDER BY s.name, p.name, i.id
    """, day=day)

    def fmt_dioptria(r):
        if r.get("sphere") is not None or r.get("cylinder") is not None:
            esf = f"{r['sphere']:+.2f}" if r.get("sphere") is not None else "-"
            cil = f"{r['cylinder']:+.2f}" if r.get("cylinder") is not None else "-"
            return f"Esf {esf} / Cil {cil}"
        else:
            b   = f"{r['base']:.2f}" if r.get("base") is not None else "-"
            add = f"+{r['addition']:.2f}" if r.get("addition") is not None else "-"
            return f"Base {b} / Adição {add}"

    return render_template("extornos.html", items=items, day=day, fmt_dioptria=fmt_dioptria)


@app.route("/extornos/<int:item_id>/criar", methods=["POST"])
def extornos_criar(item_id):
    ret = require_role("pagador","admin")
    if ret: return ret

    day = (request.args.get("date") or date.today().isoformat())

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO supplier_credits (supplier_id, item_id, amount, created_at)
            SELECT 
              o.supplier_id, 
              i.id,
              COALESCE(i.quantity,0) * COALESCE(i.unit_price,0.0),
              :now
            FROM purchase_items i
            JOIN purchase_orders o ON o.id = i.order_id
            WHERE i.id = :item_id
            ON CONFLICT (item_id) DO NOTHING
        """), dict(now=datetime.utcnow(), item_id=item_id))

    audit("extorno_create", f"item_id={item_id}")
    flash("Extorno registrado como crédito para o fornecedor.", "success")
    return redirect(url_for("extornos_index", date=day))