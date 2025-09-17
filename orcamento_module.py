# orcamento_module.py
import re
from decimal import Decimal, InvalidOperation
from sqlalchemy import text

# ===================== helpers =====================
def _to_decimal(v, default=0):
    if v is None:
        return Decimal(default)
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(default)

def _to_bool01(v):
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ("1", "sim", "s", "true", "t", "x")

def _split_codes(s):
    if not s:
        return []
    parts = re.split(r"[.;,]\s*", str(s).strip())
    return [p.strip() for p in parts if p.strip()]

def _parse_acrescimos(expr):
    """
    Aceita:
      10S
      10S[esf:-2..+2]
      10S[cil:-1..0]
      10S[esf:-2..+2; cil:-1..0]
      10S[esf:-6..+6; cil:0..-2], 30S[esf:-8..+8]
    Retorna: [{code, esf_min, esf_max, cil_min, cil_max}, ...]
    """
    if not expr:
        return []
    out = []
    items = re.split(r"[.;,]\s*", str(expr).strip())
    for it in items:
        if not it:
            continue
        m = re.match(r"^\s*([A-Za-z0-9]+)\s*(?:\[(.*?)\])?\s*$", it)
        if not m:
            continue
        code, body = m.group(1), m.group(2)
        rec = {"code": code, "esf_min": None, "esf_max": None, "cil_min": None, "cil_max": None}
        if body:
            for part in re.split(r"\s*;\s*", body):
                pm = re.match(
                    r"^(esf|cil)\s*:\s*([+\-]?\d+(?:[.,]\d+)?)\s*(?:\.\.|a|to|-)\s*([+\-]?\d+(?:[.,]\d+)?)\s*$",
                    part, flags=re.I
                )
                if pm:
                    kind = pm.group(1).lower()
                    v1 = _to_decimal(pm.group(2))
                    v2 = _to_decimal(pm.group(3))
                    lo = min(v1, v2); hi = max(v1, v2)
                    if kind == "esf":
                        rec["esf_min"], rec["esf_max"] = lo, hi
                    else:
                        rec["cil_min"], rec["cil_max"] = lo, hi
        out.append(rec)
    return out

# ===================== schema =====================
DDL_ORCAMENTO = """
CREATE TABLE IF NOT EXISTS orc_produto (
  id BIGSERIAL PRIMARY KEY,
  code TEXT UNIQUE NOT NULL,
  name TEXT NOT NULL,
  price NUMERIC(12,2) NOT NULL DEFAULT 0,
  visao TEXT NOT NULL CHECK (visao IN ('visao_simples','progressiva','bifocal')),
  ar BOOLEAN NOT NULL DEFAULT FALSE,
  foto BOOLEAN NOT NULL DEFAULT FALSE,
  azul BOOLEAN NOT NULL DEFAULT FALSE,
  esf_min NUMERIC(6,2) NOT NULL DEFAULT 0,
  esf_max NUMERIC(6,2) NOT NULL DEFAULT 0,
  cil_min NUMERIC(6,2) NOT NULL DEFAULT 0,
  cil_max NUMERIC(6,2) NOT NULL DEFAULT 0,
  updated_at TIMESTAMP WITHOUT TIME ZONE DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS orc_servico_catalogo (
  code TEXT PRIMARY KEY,
  description TEXT,
  price NUMERIC(12,2) NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS orc_produto_serv_obrig (
  id BIGSERIAL PRIMARY KEY,
  produto_id BIGINT NOT NULL REFERENCES orc_produto(id) ON DELETE CASCADE,
  serv_code TEXT NOT NULL REFERENCES orc_servico_catalogo(code) ON DELETE RESTRICT,
  UNIQUE (produto_id, serv_code)
);

CREATE TABLE IF NOT EXISTS orc_produto_serv_opc (
  id BIGSERIAL PRIMARY KEY,
  produto_id BIGINT NOT NULL REFERENCES orc_produto(id) ON DELETE CASCADE,
  serv_code TEXT NOT NULL REFERENCES orc_servico_catalogo(code) ON DELETE RESTRICT,
  UNIQUE (produto_id, serv_code)
);

CREATE TABLE IF NOT EXISTS orc_produto_acrescimo (
  id BIGSERIAL PRIMARY KEY,
  produto_id BIGINT NOT NULL REFERENCES orc_produto(id) ON DELETE CASCADE,
  serv_code TEXT NOT NULL REFERENCES orc_servico_catalogo(code) ON DELETE RESTRICT,
  esf_min NUMERIC(6,2), esf_max NUMERIC(6,2),
  cil_min NUMERIC(6,2), cil_max NUMERIC(6,2),
  UNIQUE (produto_id, serv_code, esf_min, esf_max, cil_min, cil_max)
);
"""

def ensure_orcamento_schema(engine):
    with engine.begin() as conn:
        conn.execute(text(DDL_ORCAMENTO))
        # Garante que users.role aceita 'cliente'
        try:
            conn.execute(text("""
            DO $$
            DECLARE r record;
            BEGIN
              FOR r IN
                SELECT c.conname
                FROM pg_constraint c
                JOIN pg_class t ON c.conrelid = t.oid
                WHERE t.relname='users' AND c.contype='c'
                  AND pg_get_constraintdef(c.oid) ILIKE '%role IN (%'
              LOOP
                EXECUTE format('ALTER TABLE users DROP CONSTRAINT %I', r.conname);
              END LOOP;
              BEGIN
                ALTER TABLE users
                ADD CONSTRAINT users_role_check
                CHECK (role IN ('admin','comprador','pagador','cliente'));
              EXCEPTION WHEN duplicate_object THEN
                -- já existe
              END;
            END $$;
            """))
        except Exception:
            pass

# ===================== rotas =====================
def register_orcamento(app, engine, require_role):
    from flask import render_template, request, redirect, url_for, flash

    @app.route("/orcamento", methods=["GET"])
    def orcamento():
        ret = require_role("admin", "comprador", "pagador", "cliente")
        if ret: return ret
        return render_template("orcamento.html")

    @app.post("/api/orcamento/options")
    def api_orcamento_options():
        ret = require_role("admin","comprador","pagador","cliente")
        if ret: return ret

        data = request.get_json(force=True) or {}
        visao = (data.get("visao") or "").strip()
        flags = data.get("flags") or {}
        ar   = bool(flags.get("ar"))
        foto = bool(flags.get("foto"))
        azul = bool(flags.get("azul"))

        def _nz(n):
            try: return Decimal(str(n)) if n not in (None,"") else Decimal("0")
            except Exception: return Decimal("0")

        od = data.get("od") or {}; oe = data.get("oe") or {}
        od_esf = _nz(od.get("esf")); od_cil = _nz(od.get("cil"))
        oe_esf = _nz(oe.get("esf")); oe_cil = _nz(oe.get("cil"))

        q = text("""
          SELECT id, name, price, esf_min, esf_max, cil_min, cil_max
          FROM orc_produto
          WHERE visao=:visao AND ar=:ar AND foto=:foto AND azul=:azul
        """)
        products = []
        with engine.begin() as conn:
            for r in conn.execute(q, {"visao":visao,"ar":ar,"foto":foto,"azul":azul}).mappings():
                def inrng(v, a, b):
                    lo = min(Decimal(a), Decimal(b)); hi = max(Decimal(a), Decimal(b))
                    return Decimal(v) >= lo and Decimal(v) <= hi
                ok_od = inrng(od_esf, r["esf_min"], r["esf_max"]) and inrng(od_cil, r["cil_min"], r["cil_max"])
                ok_oe = inrng(oe_esf, r["esf_min"], r["esf_max"]) and inrng(oe_cil, r["cil_min"], r["cil_max"])
                if ok_od and ok_oe:
                    products.append({"id": int(r["id"]), "name": r["name"], "price": float(r["price"] or 0)})
        return {"products": products}

    @app.post("/api/orcamento/services")
    def api_orcamento_services():
        ret = require_role("admin","comprador","pagador","cliente")
        if ret: return ret

        data = request.get_json(force=True) or {}
        pid = int(data.get("product_id"))

        def _nz(n):
            try: return Decimal(str(n)) if n not in (None,"") else Decimal("0")
            except Exception: return Decimal("0")

        od = data.get("od") or {}; oe = data.get("oe") or {}
        od_esf = _nz(od.get("esf")); od_cil = _nz(od.get("cil"))
        oe_esf = _nz(oe.get("esf")); oe_cil = _nz(oe.get("cil"))

        with engine.begin() as conn:
            ob = conn.execute(text("""
              SELECT c.code, COALESCE(c.description,c.code) AS name, COALESCE(c.price,0) AS price
              FROM orc_produto_serv_obrig o
              JOIN orc_servico_catalogo c ON c.code=o.serv_code
              WHERE o.produto_id=:pid
            """), {"pid": pid}).mappings().all()

            op = conn.execute(text("""
              SELECT c.code, COALESCE(c.description,c.code) AS name, COALESCE(c.price,0) AS price
              FROM orc_produto_serv_opc o
              JOIN orc_servico_catalogo c ON c.code=o.serv_code
              WHERE o.produto_id=:pid
            """), {"pid": pid}).mappings().all()

            ac = conn.execute(text("""
              SELECT a.serv_code, a.esf_min, a.esf_max, a.cil_min, a.cil_max,
                     COALESCE(c.description,a.serv_code) AS name, COALESCE(c.price,0) AS price
              FROM orc_produto_acrescimo a
              JOIN orc_servico_catalogo c ON c.code=a.serv_code
              WHERE a.produto_id=:pid
            """), {"pid": pid}).mappings().all()

        def within(v, a, b):
            if a is None and b is None: return False
            a = Decimal(a) if a is not None else Decimal(v)
            b = Decimal(b) if b is not None else Decimal(v)
            lo = min(a,b); hi = max(a,b)
            return Decimal(v) >= lo and Decimal(v) <= hi

        mandatory = [{"id": r["code"], "name": r["name"], "price": float(r["price"])} for r in ob]

        for a in ac:
            trig_od = within(od_esf, a["esf_min"], a["esf_max"]) and within(od_cil, a["cil_min"], a["cil_max"])
            trig_oe = within(oe_esf, a["esf_min"], a["esf_max"]) and within(oe_cil, a["cil_min"], a["cil_max"])
            if trig_od or trig_oe:
                mandatory.append({"id": a["serv_code"], "name": a["name"], "price": float(a["price"])})

        optional = [{"id": r["code"], "name": r["name"], "price": float(r["price"])} for r in op]
        return {"mandatory": mandatory, "optional": optional}

    # -------- Importar regras do orçamento (Admin) --------
    @app.post("/admin/import_orcamento")
    def admin_import_orcamento():
        ret = require_role("admin")
        if ret: return ret

        from flask import render_template, request, redirect, url_for, flash
        from openpyxl import load_workbook

        file = request.files.get("file_orcamento")
        if not file or file.filename == "":
            flash("Selecione um arquivo .xlsx", "error")
            return redirect(request.referrer or url_for("admin_import"))

        try:
            wb = load_workbook(filename=file, data_only=True)
            ws = wb.worksheets[0]
        except Exception as e:
            flash(f"Erro ao ler Excel: {e}", "error")
            return redirect(request.referrer or url_for("admin_import"))

        # headers
        header_map = {}
        for j, c in enumerate(ws[1], start=1):
            header_map[(str(c.value or "").strip().lower())] = j

        def col(*names):
            for n in names:
                n = n.lower()
                if n in header_map: return header_map[n]
            return None

        idx_prod   = col("produto","nome")
        idx_code   = col("código","codigo","cod")
        idx_valor  = col("valor","preço","preco")
        idx_visao  = col("tipo de visão","tipo de visao","visao","visão","tv")
        idx_ar     = col("antirreflexo","anti-reflexo","ar")
        idx_foto   = col("fotosensível","fotossensivel","foto")
        idx_azul   = col("filtro azul","azul","blue")
        idx_esfmin = col("esf mínimo","esf minimo","esf min","esf_min")
        idx_esfmax = col("esf máximo","esf maximo","esf max","esf_max")
        idx_cilmin = col("cil mínimo","cil minimo","cil min","cil_min")
        idx_cilmax = col("cil máximo","cil maximo","cil max","cil_max")
        idx_servob = col("serviços obrigatórios","servicos obrigatorios","obrigatórios","obrigatorios","n1-14")
        idx_servop = col("serviços disponíveis","servicos disponiveis","disponíveis","disponiveis","o15")
        idx_acresc = col("acréscimos","acrescimos","acréscimo","acrescimo")

        must = [idx_prod,idx_code,idx_valor,idx_visao,idx_ar,idx_foto,idx_azul,idx_esfmin,idx_esfmax,idx_cilmin,idx_cilmax,idx_servob,idx_servop]
        if not all(must):
            flash("Cabeçalhos ausentes na 1ª aba. Confira os nomes.", "error")
            return redirect(request.referrer or url_for("admin_import"))

        prod_ins = prod_upd = 0
        serv_ob_upserts = serv_opc_upserts = acresc_upserts = 0
        rows = 0

        with engine.begin() as conn:
            for i in range(2, ws.max_row + 1):
                rows += 1
                nome = (ws.cell(i, idx_prod).value or "").strip()
                code = (str(ws.cell(i, idx_code).value or "").strip())
                if not nome or not code: continue

                vis_raw = (str(ws.cell(i, idx_visao).value or "").strip().lower())
                if vis_raw in ("vs","visao simples","visão simples","visao_simples","visão_simples"):
                    visao = "visao_simples"
                elif vis_raw in ("progressiva",):
                    visao = "progressiva"
                elif vis_raw in ("bifocal",):
                    visao = "bifocal"
                else:
                    visao = "visao_simples"

                price  = _to_decimal(ws.cell(i, idx_valor).value, 0)
                ar     = _to_bool01(ws.cell(i, idx_ar).value)
                foto   = _to_bool01(ws.cell(i, idx_foto).value)
                azul   = _to_bool01(ws.cell(i, idx_azul).value)
                esfmin = _to_decimal(ws.cell(i, idx_esfmin).value, 0)
                esfmax = _to_decimal(ws.cell(i, idx_esfmax).value, 0)
                cilmin = _to_decimal(ws.cell(i, idx_cilmin).value, 0)
                cilmax = _to_decimal(ws.cell(i, idx_cilmax).value, 0)

                r = conn.execute(text("""
                  INSERT INTO orc_produto (code,name,price,visao,ar,foto,azul,esf_min,esf_max,cil_min,cil_max,updated_at)
                  VALUES (:code,:name,:price,:visao,:ar,:foto,:azul,:esfmin,:esfmax,:cilmin,:cilmax,NOW())
                  ON CONFLICT (code) DO UPDATE SET
                    name=EXCLUDED.name, price=EXCLUDED.price, visao=EXCLUDED.visao,
                    ar=EXCLUDED.ar, foto=EXCLUDED.foto, azul=EXCLUDED.azul,
                    esf_min=EXCLUDED.esf_min, esf_max=EXCLUDED.esf_max,
                    cil_min=EXCLUDED.cil_min, cil_max=EXCLUDED.cil_max,
                    updated_at=NOW()
                  RETURNING (xmax = 0) AS inserted, id
                """), {"code":code,"name":nome,"price":price,"visao":visao,"ar":ar,"foto":foto,"azul":azul,
                       "esfmin":esfmin,"esfmax":esfmax,"cilmin":cilmin,"cilmax":cilmax})
                row = r.fetchone()
                inserted = bool(row[0]); pid = row[1]
                prod_ins += 1 if inserted else 0
                prod_upd += 0 if inserted else 1

                for sc in _split_codes(ws.cell(i, idx_servob).value):
                    conn.execute(text("INSERT INTO orc_servico_catalogo (code) VALUES (:c) ON CONFLICT DO NOTHING"), {"c": sc})
                    conn.execute(text("""
                      INSERT INTO orc_produto_serv_obrig (produto_id, serv_code)
                      VALUES (:pid, :sc)
                      ON CONFLICT (produto_id, serv_code) DO NOTHING
                    """), {"pid": pid, "sc": sc})
                    serv_ob_upserts += 1

                for sc in _split_codes(ws.cell(i, idx_servop).value):
                    conn.execute(text("INSERT INTO orc_servico_catalogo (code) VALUES (:c) ON CONFLICT DO NOTHING"), {"c": sc})
                    conn.execute(text("""
                      INSERT INTO orc_produto_serv_opc (produto_id, serv_code)
                      VALUES (:pid, :sc)
                      ON CONFLICT (produto_id, serv_code) DO NOTHING
                    """), {"pid": pid, "sc": sc})
                    serv_opc_upserts += 1

                if idx_acresc:
                    for ac in _parse_acrescimos(ws.cell(i, idx_acresc).value):
                        conn.execute(text("INSERT INTO orc_servico_catalogo (code) VALUES (:c) ON CONFLICT DO NOTHING"),
                                     {"c": ac["code"]})
                        conn.execute(text("""
                          INSERT INTO orc_produto_acrescimo (produto_id, serv_code, esf_min, esf_max, cil_min, cil_max)
                          VALUES (:pid,:sc,:esfmin,:esfmax,:cilmin,:cilmax)
                          ON CONFLICT (produto_id, serv_code, esf_min, esf_max, cil_min, cil_max) DO NOTHING
                        """), {"pid":pid,"sc":ac["code"],
                               "esfmin":ac["esf_min"],"esfmax":ac["esf_max"],
                               "cilmin":ac["cil_min"],"cilmax":ac["cil_max"]})
                        acresc_upserts += 1

        return render_template("admin_import.html", imp_orcamento={
            "prod_inserted": prod_ins, "prod_updated": prod_upd,
            "serv_obrig_upserts": serv_ob_upserts, "serv_opc_upserts": serv_opc_upserts,
            "acresc_upserts": acresc_upserts, "rows": rows
        })
